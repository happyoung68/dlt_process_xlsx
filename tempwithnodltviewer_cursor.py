import base64
import sys
from PyQt5.QtWidgets import (
    QPushButton, QMessageBox, QFileDialog, QProgressBar, QTextEdit,
    QDateTimeEdit, QTableWidget, QTableWidgetItem, QSpinBox, QHeaderView,
    QMainWindow, QLabel, QLineEdit, QApplication, QWidget, QVBoxLayout, QHBoxLayout, QGroupBox
)
from PyQt5.QtCore import QSettings, QThread, pyqtSignal, QDateTime
from PyQt5.QtCore import Qt
import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
import gzip
import shutil
import time
import io
import tempfile
import tarfile
import struct
import re

# ===================== 嵌入DLT原生解析器（无依赖、高性能版） =====================
class LogParser:
    """DLT日志解析器（适配non-verbose模式，无第三方依赖）"""
    MAGIC_HEADER = b'DLT\x01'
    STORAGE_HEADER_LEN = 16
    STANDARD_HEADER_LEN = 16
    EXTENDED_HEADER_LEN = 10
    MIN_MSG_LENGTH = 42

    MESSAGE_TYPE_MAP = {
        0x00: "LOG",
        0x01: "APP_TRACE",
        0x02: "NW_TRACE",
        0x03: "CONTROL"
    }
    LOG_LEVEL_MAP = {
        0x01: "FATAL",
        0x02: "ERROR",
        0x03: "WARN",
        0x04: "INFO",
        0x05: "DEBUG",
        0x06: "VERBOSE"
    }
    DATA_SIZE_MAP = {0x01: 1, 0x02: 2, 0x03: 4, 0x04: 8}
    ENCODING_MAP = {0x00: 'ascii', 0x01: 'utf-8', 0x02: 'hex_ascii', 0x04: 'utf-8'}
    
    STORAGE_HEADER_FMT = '<Ii'
    LENGTH_FMT = '>H'
    SESSION_ID_FMT = '>I'
    TIMESTAMP_FMT = '>I'
    STR_LEN_FMT = '<H'
    NON_VERBOSE_MSG_ID_LEN = 4
    DLT_SEPARATOR = b'DLT\x01'
    CAL_TIMESTAMP_PATTERN = re.compile(r'\[\#:(\d+\.\d+)s\]')

    @classmethod
    def parse_storage_header(cls, hex_data):
        if len(hex_data) < cls.STORAGE_HEADER_LEN:
            raise ValueError("StorageHeader长度不足16字节")
        if hex_data[:4] != cls.MAGIC_HEADER:
            raise ValueError("无效的DLT文件格式")

        seconds, microseconds = struct.unpack(cls.STORAGE_HEADER_FMT, hex_data[4:12])
        ecuid = hex_data[12:16].decode('ascii', errors='ignore')
        unix_ts = seconds + microseconds / 1000000.0
        abs_time = datetime.fromtimestamp(unix_ts)
        time_str = f"{abs_time.strftime('%Y/%m/%d %H:%M:%S.')}{microseconds:06d}"

        return {
            'magic': cls.MAGIC_HEADER.hex(),
            'seconds': seconds,
            'microseconds': microseconds,
            'ecuid': ecuid,
            'time': time_str,
            'unix_ts': unix_ts,
        }

    @classmethod
    def parse_standard_header(cls, hex_data):
        if len(hex_data) < cls.STANDARD_HEADER_LEN:
            raise ValueError("StandardHeader长度不足16字节")

        header_type_byte = hex_data[0]
        use_extended_header = (header_type_byte & 0x01) != 0
        msb_first = (header_type_byte & 0x02) != 0
        with_ecuid = (header_type_byte & 0x04) != 0
        with_session_id = (header_type_byte & 0x08) != 0
        with_timestamp = (header_type_byte & 0x10) != 0
        version_number = (header_type_byte >> 5) & 0x07

        message_counter = hex_data[1]
        length = struct.unpack(cls.LENGTH_FMT, hex_data[2:4])[0]

        offset = 4
        ecuid = ""
        if with_ecuid and offset + 4 <= len(hex_data):
            ecuid = hex_data[offset:offset+4].decode('ascii', errors='ignore')
            offset += 4

        session_id = 0
        if with_session_id and offset + 4 <= len(hex_data):
            session_id = struct.unpack(cls.SESSION_ID_FMT, hex_data[offset:offset+4])[0]
            offset += 4

        timestamp = 0.0
        if with_timestamp and offset + 4 <= len(hex_data):
            timestamp = struct.unpack(cls.TIMESTAMP_FMT, hex_data[offset:offset+4])[0] / 10000.0

        return {'header_type': header_type_byte, 'use_extended_header': use_extended_header, 'msb_first': msb_first,
                'with_ecuid': with_ecuid, 'with_session_id': with_session_id, 'with_timestamp': with_timestamp,
                'version_number': version_number, 'message_counter': message_counter, 'length': length,
                'ecuid': ecuid, 'session_id': session_id, 'timestamp': timestamp}

    @classmethod
    def parse_extended_header(cls, hex_data):
        if len(hex_data) < cls.EXTENDED_HEADER_LEN:
            raise ValueError("ExtendedHeader长度不足10字节")

        message_info = hex_data[0]
        mode = (message_info & 0x01) != 0
        message_type_bits = (message_info >> 1) & 0x07
        message_type_info_bits = (message_info >> 4) & 0x0F

        message_type = cls.MESSAGE_TYPE_MAP.get(message_type_bits, f"RESERVED_{message_type_bits}")
        message_type_info = cls.LOG_LEVEL_MAP.get(message_type_info_bits, "") if message_type == "LOG" else ""

        num_arguments = hex_data[1]
        application_id = hex_data[2:6].decode('ascii', errors='ignore').replace('\x00', '')
        context_id = hex_data[6:10].decode('ascii', errors='ignore').replace('\x00', '')

        return {'message_info': message_info, 'mode': mode, 'message_type': message_type,
                'message_type_info': message_type_info, 'num_arguments': num_arguments,
                'application_id': application_id, 'context_id': context_id}

    @classmethod
    def parse_payload(cls, hex_data, num_arguments, msb_first, mode):
        if not mode:
            return cls._parse_non_verbose_payload(hex_data)

        payload_parts = []
        payload_p_parts = []
        offset = 0
        total_length = len(hex_data)
        data_size_map = cls.DATA_SIZE_MAP
        encoding_map = cls.ENCODING_MAP
        str_len_fmt = cls.STR_LEN_FMT
        max_offset = total_length - 4

        for _arg_idx in range(num_arguments):
            if offset > max_offset:
                break

            type_info = int.from_bytes(hex_data[offset:offset+4], 'little')
            offset += 4

            data_type_length = type_info & 0x0F
            is_bool = (type_info & 0x10) != 0
            is_sint = (type_info & 0x20) != 0
            is_usint = (type_info & 0x40) != 0
            is_float = (type_info & 0x80) != 0
            is_string = (type_info & 0x200) != 0
            is_raw = (type_info & 0x400) != 0
            string_code = (type_info >> 15) & 0x07

            result = ""
            group_part = ""
            is_special_string = False

            if string_code == 0x04:
                is_string = True
                is_special_string = True
                num_arguments = 1

            if is_string:
                if is_special_string:
                    if offset + 2 > total_length:
                        result = ""
                    else:
                        str_len = struct.unpack('>H', hex_data[offset:offset+2])[0]
                        offset += 2
                        if str_len <= 0 or offset + str_len > total_length:
                            result = ""
                        else:
                            string_data = hex_data[offset:offset+str_len]
                            offset += str_len
                            encoding = encoding_map.get(string_code, 'ascii')
                            result = string_data.decode(encoding, errors='ignore')
                            group_part = result
                    payload_parts.append(result)
                    payload_p_parts.append(group_part)
                    break
                else:
                    if offset + 2 > total_length:
                        break
                    str_len = struct.unpack(str_len_fmt, hex_data[offset:offset+2])[0]
                    offset += 2
                    if offset + str_len > total_length or str_len == 0:
                        break

                    string_data = hex_data[offset:offset+str_len]
                    offset += str_len
                    encoding = encoding_map.get(string_code, 'ascii')

                    if encoding == 'hex_ascii':
                        hex_str = string_data.hex() if msb_first else string_data[::-1].hex()
                        result = f"0x{hex_str.upper()}"
                    else:
                        result = string_data.decode(encoding, errors='ignore')
                        group_part = result

            elif is_raw:
                if offset + 2 > total_length:
                    break
                raw_len = struct.unpack(str_len_fmt, hex_data[offset:offset+2])[0]
                offset += 2
                if offset + raw_len > total_length or raw_len == 0:
                    break

                raw_data = hex_data[offset:offset+raw_len]
                offset += raw_len
                result = raw_data.hex(' ', 1)
                group_part = f"{{arg.{(_arg_idx + 1):02d}, raw}}"

            elif is_sint or is_usint or is_float or is_bool:
                data_size = data_size_map.get(data_type_length, 0)
                if data_size == 0 or offset + data_size > total_length:
                    break

                data_bytes = hex_data[offset:offset+data_size]
                offset += data_size
                fmt_prefix = '>' if msb_first else '<'
                fmt = ""

                if is_sint:
                    fmt = f"{fmt_prefix}{'bhi'[data_size // 2]}" if data_size in (1,2,4) else ""
                elif is_usint:
                    fmt = f"{fmt_prefix}{'BHI'[data_size // 2]}" if data_size in (1,2,4) else ""

                is_hex_format = False
                if fmt:
                    try:
                        result = str(struct.unpack(fmt, data_bytes)[0])
                    except struct.error:
                        result = ""

                if is_usint and string_code == 0x02 and result:
                    hex_str = data_bytes.hex() if msb_first else data_bytes[::-1].hex()
                    result = f"0x{hex_str.upper()}"
                    is_hex_format = True

                if is_bool:
                    datatype = 'bool'
                elif is_sint:
                    datatype = f"sint{ {1:'8',2:'16',4:'32',8:'64'}[data_size] }"
                elif is_usint:
                    datatype = f"uint{ {1:'8',2:'16',4:'32',8:'64'}[data_size] }"
                    if is_hex_format:
                        datatype += "_hexstr"
                elif is_float:
                    datatype = f"float{ {4:'32',8:'64'}[data_size] }"
                else:
                    datatype = ''

                group_part = f"{{arg.{(_arg_idx + 1):02d}, {datatype}}}" if datatype else result
            else:
                result = hex_data[offset:].hex(' ', 1) if offset < total_length else ""
                offset = total_length
                group_part = result

            payload_parts.append(result)
            payload_p_parts.append(group_part)

        payload = " ".join(payload_parts).replace('\x00', '').strip()
        payload_p = " ".join(payload_p_parts).replace('\x00', '').strip()
        return payload, payload_p

    @classmethod
    def _parse_non_verbose_payload(cls, hex_data):
        try:
            if len(hex_data) < cls.NON_VERBOSE_MSG_ID_LEN:
                message_id_str = "[0]"
                data_bytes = b""
            else:
                message_id = struct.unpack('>I', hex_data[:cls.NON_VERBOSE_MSG_ID_LEN])[0]
                message_id_str = f"[{message_id}]"
                data_bytes = hex_data[cls.NON_VERBOSE_MSG_ID_LEN:]
                separator_idx = data_bytes.find(cls.DLT_SEPARATOR)
                if separator_idx != -1:
                    data_bytes = data_bytes[:separator_idx]

            if not data_bytes:
                data_part = ""
            else:
                ascii_str = data_bytes.decode('ascii', errors='ignore')
                hex_str = data_bytes.hex(' ', 1)
                data_part = f"{ascii_str}|{hex_str}"

            payload = f"{message_id_str} {data_part}".strip()
            return payload, payload
        except Exception:
            return "", ""

    @classmethod
    def split_dlt_messages(cls, file_data):
        messages = []
        start = 0
        magic = cls.MAGIC_HEADER
        magic_len = len(magic)
        file_len = len(file_data)

        while start < file_len:
            next_start = file_data.find(magic, start + magic_len)
            if next_start == -1:
                messages.append(file_data[start:])
                break
            messages.append(file_data[start:next_start])
            start = next_start
        return messages

    @classmethod
    def _extract_cal_timestamp(cls, ecuid, payload, timestamp):
        try:
            if ecuid in ["CCU0", "CCU1"]:
                match = cls.CAL_TIMESTAMP_PATTERN.search(payload)
                if match:
                    return round(float(match.group(1)), 4)
            return round(timestamp, 4) if timestamp is not None else 0.0
        except Exception:
            return round(timestamp, 4) if timestamp is not None else 0.0

    @classmethod
    def parse_dlt_message(cls, hex_data, msg_index):
        if len(hex_data) < cls.MIN_MSG_LENGTH:
            return None
        try:
            storage_header = cls.parse_storage_header(hex_data[:cls.STORAGE_HEADER_LEN])
            standard_header = cls.parse_standard_header(hex_data[cls.STORAGE_HEADER_LEN:cls.STORAGE_HEADER_LEN+cls.STANDARD_HEADER_LEN])
            payload_length = standard_header['length'] - cls.STANDARD_HEADER_LEN - cls.EXTENDED_HEADER_LEN
            
            if payload_length < 0:
                return None

            extended_header = cls.parse_extended_header(hex_data[32:42])
            payload_data = hex_data[42:42+payload_length]
            payload, payload_p = cls.parse_payload(payload_data, extended_header['num_arguments'], standard_header['msb_first'], extended_header['mode'])
            cal_timestamp = cls._extract_cal_timestamp(storage_header['ecuid'], payload, standard_header['timestamp'])
            unix_ts = storage_header.get('unix_ts')

            return (
                msg_index, storage_header['time'], standard_header['timestamp'], cal_timestamp,
                standard_header['message_counter'], storage_header['ecuid'],
                extended_header['application_id'], extended_header['context_id'],
                standard_header['session_id'], extended_header['message_type'],
                extended_header['message_type_info'], 'verbose' if extended_header['mode'] else 'non-verbose',
                extended_header['num_arguments'], payload, payload_p
                , unix_ts
            )
        except Exception:
            return None

# ===================== 原有业务逻辑（保留不变） =====================
RT1_signals = [
    ["", "", "", "NMiDPSNMKeepAwakeFlag", "NMEEMKeepAwakeFlag", "NMEEMKeepAwakeFlag_Aux",
     "NMVCUGlobalKeepAwakeFlag", "NMVCUChargingKeepAwakeFlag(新能源,VCU_hcu)"],
    ["DoIPActiveLineStatus", "NMBMSGlobalKeepAwakeFlag", "NMBMSChargingKeepAwakeFlag",
     "NMBMSThermalRunawayKeepAwakeFlag(新能源,BMS)", "-", "NMBMSThermalRunawayKeepAwakeFlag(动力,VCU_slccu)",
     "NMVCUChargingKeepAwakeFlag(动力,VCU_slccu)", "NMVCU2GlobalKeepAwakeFlag"],
    ["NMVHMKeepAwakeFlag", ".", "CSM_RC_KeepAwake", "NMHVACKeepAwakeFlag(新能源,TMS)", ".", ".",
     "BDC_NMBDCKeepAwakeFlag", "NMTMSChargingKeepAwakeFlag"],
    ["NMTMSGlobalKeepAwakeFlag", "NMHMPNCDownloadKeepAwakeFlag", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    ["NMWCCDownloadRequest_RT1", "CAN7_DiagMSG_ID ！= 0x74F（PNCGlobal）", "CAN7_DiagMSG_ID == 0x74F（RequestOBD）", ".",
     ".", ".", ".", "."],
    ["SPME触发A核网络维持", "SMM触发A核网络维持", "CSM触发A核网络维持", "OTACLT触发A核网络维持", "VCP触发A核网络维持",
     "BDC触发A核网络维持", "VT触发A核网络维持", "."],
    [".", ".", ".", ".", ".", ".", ".", "."]
]

Event_Source_signals = ["PNCGlobal", "PNCCharging", "PNCBSM", "PNCDownload", "PNCEnter", "PNCHazard", "Reserved", "PNCRearlight"]
Acore_signals = ["SPME(PNCGlobal)", "SMM(PNCGlobal)", "CSM(PNCGlobal)", "OTACLT(PNCDownload)", "VCP(PNCGlobal)",
                 "BDC(PNCDownload,PNCGlobal)", "VT(PNCGlobal)", "UA"]

RT2_signals = [
    ["NMATWSKeepAwakeFlag", "NMExteriorLightKeepAwakeFlag", "NMLockingKeepAwakeFlag", "PWL_NMPWLKeepAwakeFlag",
     "SR_NMSunRoofKeepAwakeFlag", "NMPLGKeepAwakeFlag", "NMPowerOperatedDoorKeepAwakeFlag", "MR_NMMirrorKeepAwakeFlag"],
    ["SSW_NMSeatControlKeepAwakeFlag", "NMHVACKeepAwakeFlag", "NMHVACRLSLinAwakeFlag", "NMAQCKeepAwakeFlag",
     "NMTPMSKeepAwakeFlag", "ILC_NMKeepAwakeFlag", "NMESMKeepAwakeFlag", "NMKAEKeepAwakeFlag"],
    ["NMVRMPNCCharingKeepAwakeFlag", "SMM_NMLPDKeepAwakeFlag", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    ["NMWCCDownloadRequest_RT2", ".", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."]
]

def rt1_detect_and_modify(txt_file):
    keyword = ["RT1_McoreSignalChange", "RT1_NetWUStVar"]
    df = pd.read_csv(txt_file, delimiter='\s+')
    ninth_column = df.iloc[:, 13]
    keyword_indices = ninth_column[ninth_column.str.contains('|'.join(keyword))].index
    for index in keyword_indices:
        ans = "--"
        temp_str = df.iloc[index, 13]
        numbers_string = temp_str.split("):")[1]
        numbers_list = numbers_string.split(",")[:8]
        bin_array = []
        for dec in numbers_list:
            decimal_value = int(dec, 16)
            binary_string = bin(decimal_value)[2:]
            padded_binary_str = binary_string.zfill(8)
            bin_array.append(padded_binary_str)
        car_state = bin_array[0][5] + bin_array[0][6] + bin_array[0][7]
        if car_state == "000":
            ans += "Standby--"
        elif car_state == "001":
            ans += "Comfort--"
        elif car_state == "011":
            ans += "DRV--"
        for i in range(len(bin_array)):
            for j in range(8):
                if bin_array[i][j] == "1":
                    ans += RT1_signals[i][7 - j]
                    ans += "--"
        df.iloc[index, 13] += f"{ans}"
    df.to_csv(txt_file, sep=' ', index=False)

def rt2_detect_and_modify(txt_file):
    try:
        keyword = ["RT2_McoreSignalChange", "RT2_NetWUStVar"]
        df = pd.read_csv(txt_file, delimiter='\s+', dtype=str)
        ninth_column = df.iloc[:, 13]
        keyword_indices = ninth_column[ninth_column.str.contains('|'.join(keyword))].index
        for index in keyword_indices:
            ans = "--"
            temp_str = df.iloc[index, 13]
            numbers_string = temp_str.split("):")[1]
            numbers_list = numbers_string.split(",")[:8]
            bin_array = []
            for dec in numbers_list:
                decimal_value = int(dec, 16)
                binary_string = bin(decimal_value)[2:]
                padded_binary_str = binary_string.zfill(8)
                bin_array.append(padded_binary_str)
            for i in range(len(bin_array)):
                for j in range(8):
                    if bin_array[i][j] == "1":
                        ans += RT2_signals[i][7 - j]
                        ans += "--"
            df.iloc[index, 13] += f"{ans}"
        df.to_csv(txt_file, sep=' ', index=False)
    except Exception as e:
        print(f"处理文件 {txt_file} 时出现错误：{str(e)}")

def eventSource_current_detect_and_modify(txt_file):
    keyword = ["EventSource_Current", "EvenSource_Current"]
    df = pd.read_csv(txt_file, delimiter='\s+')
    ninth_column = df.iloc[:, 13]
    keyword_indices = ninth_column[ninth_column.str.contains('|'.join(keyword))].index
    for index in keyword_indices:
        ans = "--"
        temp_str = df.iloc[index, 13]
        numbers_string = temp_str.split("):")[1]
        numbers_list = numbers_string.split(",")[:1]
        bin_array = []
        for dec in numbers_list:
            decimal_value = int(dec, 16)
            binary_string = bin(decimal_value)[2:]
            padded_binary_str = binary_string.zfill(8)
            bin_array.append(padded_binary_str)
        for i in range(len(bin_array)):
            for j in range(8):
                if bin_array[i][j] == "1":
                    ans += Event_Source_signals[7 - j]
                    ans += "--"
        df.iloc[index, 13] += f"{ans}"
    df.to_csv(txt_file, sep=' ', index=False)

def acore_signal_detect_and_modify(txt_file):
    keyword = "AcoreSignalChange"
    df = pd.read_csv(txt_file, delimiter='\s+')
    ninth_column = df.iloc[:, 13]
    keyword_indices = ninth_column[ninth_column.str.contains(keyword)].index
    for index in keyword_indices:
        ans = "--"
        temp_str = df.iloc[index, 13]
        numbers_string = temp_str.split("):")[1]
        numbers_list = numbers_string.split(",")[:1]
        bin_array = []
        for dec in numbers_list:
            decimal_value = int(dec, 16)
            binary_string = bin(decimal_value)[2:]
            padded_binary_str = binary_string.zfill(8)
            bin_array.append(padded_binary_str)
        for i in range(len(bin_array)):
            for j in range(8):
                if bin_array[i][j] == "1":
                    ans += Acore_signals[7 - j]
                    ans += "--"
        df.iloc[index, 13] += f"{ans}"
    df.to_csv(txt_file, sep=' ', index=False)

def ensure_dlt_extension(filename):
    if '.' not in filename:
        new_filename = filename + '.dlt'
    else:
        base_name, extension = os.path.splitext(filename)
        if extension != '.dlt':
            new_filename = base_name + '.dlt'
        else:
            new_filename = filename
    if new_filename == filename:
        return filename
    try:
        os.rename(filename, new_filename)
        return new_filename
    except OSError as e:
        print(f'Error: {e}')
        return filename

# ===================== 替换：原生DLT转TXT（无exe依赖、格式严格匹配） =====================
def convert_dlt_to_txt(input_dlt_file, output_txt_file):
    """纯Python原生DLT解析，无依赖、高性能"""
    try:
        with open(input_dlt_file, 'rb') as f:
            file_data = f.read()
        
        messages = LogParser.split_dlt_messages(file_data)
        unix_ts_list = []

        # 写入：无表头 + 单空格分隔（流式写，避免 parsed_results 全量常驻内存）
        with open(output_txt_file, 'w', encoding='utf-8') as f_out:
            for idx, msg in enumerate(messages):
                res = LogParser.parse_dlt_message(msg, idx)
                if not res:
                    continue

                # 舍弃第4列(索引3)、第15列(索引14)
                # parse_dlt_message 在末尾额外追加了 unix_ts，因此 res[4:14] 与历史输出保持一致
                filtered_row = list(res[:3]) + list(res[4:14])
                unix_ts_list.append(res[-1])

                row_str = " ".join(
                    [str(item).replace('\t', ' ').replace('\x00', '').strip() for item in filtered_row]
                )
                f_out.write(row_str + "\n")
        
        print(f"DLT文件 {input_dlt_file} 原生转换为TXT成功！")
        return unix_ts_list
    except Exception as e:
        print(f"DLT转换失败：{e}")
        return None

def txt_to_excel(txt_file, excel_file):
    df = pd.read_csv(txt_file, sep=' ', header=None)
    df.to_excel(excel_file, index=False, header=False, engine='openpyxl')
    print("转换完成，Excel 文件保存在:", excel_file)
    return df

def decompress_folder(folder):
    temp_folder = os.path.join(folder, 'temp')
    if not os.path.exists(temp_folder):
        os.makedirs(temp_folder)

    for file in os.listdir(folder):
        if file.endswith('.tar.gz'):
            tar_gz_file_path = os.path.join(folder, file)
            decompressed_folder_path = os.path.join(temp_folder, file[:-7])
            try:
                with tarfile.open(tar_gz_file_path, 'r:gz') as tar:
                    tar.extractall(path=decompressed_folder_path)
                os.remove(tar_gz_file_path)
                print(f"已解压并删除: {tar_gz_file_path}")
            except Exception as e:
                print(f"处理文件 {tar_gz_file_path} 失败: {e}")

    for root, dirs, files in os.walk(folder, topdown=False):
        for file in files:
            file_path = os.path.join(root, file)
            if 'WCC' in file or 'POWM' in file:
                target_path = os.path.join(folder, file)
                shutil.move(file_path, target_path)
                print(f"保留文件: {file_path} 移动到: {target_path}")
            else:
                os.remove(file_path)
                print(f"删除文件: {file_path}")
        for dir in dirs:
            dir_path = os.path.join(root, dir)
            os.rmdir(dir_path)
            print(f"删除文件夹: {dir_path}")

    files = os.listdir(folder)
    for file in files:
        if file.endswith('.gz'):
            gzipped_file = os.path.join(folder, file)
            output_file = os.path.join(folder, os.path.splitext(file)[0])
            if os.path.exists(output_file + ".dlt"):
                print(f"已存在解压后的文件: {output_file}")
                continue
            with gzip.open(gzipped_file, 'rb') as f_in:
                with open(output_file, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            print(f"解压缩 {gzipped_file} 到 {output_file}")

class StreamToTextEdit(io.StringIO):
    def __init__(self, text_edit):
        super().__init__()
        self.text_edit = text_edit
    def write(self, text):
        self.text_edit.insertPlainText(text)
        self.text_edit.ensureCursorVisible()

# ===================== 优化工作线程（移除exe依赖） =====================
class MyWorkerThread(QThread):
    progress_signal = pyqtSignal(int)
    finished_signal = pyqtSignal()
    records_ready_signal = pyqtSignal(list)

    def __init__(self, folder_path, parent=None):
        super().__init__(parent)
        self.folder_path = folder_path

    def run(self):
        start_time = time.time()
        print(f"程序开始于:🕑 {time.ctime(start_time)}")
        workbook = openpyxl.Workbook()
        folder_path = self.folder_path
        current_step = 0
        total_steps = 11

        decompress_folder(folder_path)
        files = os.listdir(folder_path)
        input_dlt_file = [file for file in files if file.endswith('.dlt') or file.endswith('.log') or '.' not in file]
        for i in range(0, len(input_dlt_file)):
            input_dlt_file[i] = ensure_dlt_extension(folder_path + "/" + input_dlt_file[i])
        output_txt_file = [f"{os.path.splitext(file)[0]}.txt" for file in input_dlt_file]
        all_output_txt_file = folder_path + "/AAAAAAAAALLLL.txt"
        unix_ts_all = []

        now = datetime.now()
        time_string = now.strftime("%Y-%m-%d_%H:%M:%S").replace(" ", "_").replace(":", "_")
        output_excel = folder_path + "/" + time_string + ".xlsx"
        workbook.save(output_excel)

        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        # 原生转换，无exe依赖
        for dlt, txt in zip(input_dlt_file, output_txt_file):
            unix_ts_part = convert_dlt_to_txt(dlt, txt)
            if unix_ts_part:
                unix_ts_all.extend(unix_ts_part)

        # 合并TXT
        try:
            with open(all_output_txt_file, 'w', encoding='utf-8') as outfile:
                for file_name in output_txt_file:
                    with open(file_name, 'r', encoding='utf-8') as infile:
                        outfile.write(infile.read())
            print(f"合并完成，结果保存在 {all_output_txt_file}")
        except Exception as e:
            print("合并过程中出现错误:", e)

        # 删除临时文件
        for file_path in output_txt_file:
            if os.path.exists(file_path):
                os.remove(file_path)

        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        # 后续原有逻辑（保留不变）：把多次 readlines/writelines 合并成一次流式处理
        tmp_txt_file = all_output_txt_file + ".tmp"
        with open(all_output_txt_file, 'r', encoding='utf-8') as infile, open(tmp_txt_file, 'w', encoding='utf-8') as outfile:
            for line in infile:
                # Step 1: DPT_VER/DSPEC_VER 替换（基于 token 第 14 列）
                columns = line.strip().split()
                if len(columns) >= 14:
                    if "DPT_VER" in columns[13]:
                        columns[13] = "DPT_VER_illegal"
                    elif "DSPEC_VER" in columns[13]:
                        columns[13] = "DSPEC_VER_illegal"
                line2 = ' '.join(columns)

                # Step 2: 仅把 payload 部分的空格替换为 '_'（从第 13 个空格后开始）
                space_count = 0
                index = 0
                for j, char in enumerate(line2):
                    if char == ' ':
                        space_count += 1
                        if space_count == 13:
                            index = j + 1
                            break
                if index > 0:
                    line2 = line2[:index] + line2[index:].replace(' ', '_')

                # Step 3: 合并 date/time token（保持原有逻辑）
                parts = line2.strip().split()
                if len(parts) >= 3:
                    parts[1] = parts[1] + '__' + parts[2]
                    line2 = ' '.join(parts)

                outfile.write(line2 + "\n")

        os.replace(tmp_txt_file, all_output_txt_file)

        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        rt1_detect_and_modify(all_output_txt_file)
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))
        rt2_detect_and_modify(all_output_txt_file)
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))
        eventSource_current_detect_and_modify(all_output_txt_file)
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))
        acore_signal_detect_and_modify(all_output_txt_file)
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        df_for_records = txt_to_excel(all_output_txt_file, output_excel)
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        # 为 UI 生成 records（只保留：时间/相对时间/ecuid/application_id/payload；unix_ts 用于筛选/排序）
        records = []
        try:
            n = min(len(unix_ts_all), len(df_for_records))
            if df_for_records.shape[1] >= 14:
                for i in range(n):
                    ut = unix_ts_all[i]
                    if ut is None:
                        continue
                    abs_time = datetime.fromtimestamp(float(ut)).strftime("%Y/%m/%d %H:%M:%S.%f")[:-3]
                    rel_ts = df_for_records.iloc[i, 3]
                    ecuid = str(df_for_records.iloc[i, 5])
                    app_id = str(df_for_records.iloc[i, 6])
                    payload = str(df_for_records.iloc[i, 13])
                    records.append({
                        "unix_ts": float(ut),
                        "abs_time": abs_time,
                        "rel_time": rel_ts,
                        "ecuid": ecuid,
                        "application_id": app_id,
                        "payload": payload,
                    })
        except Exception as e:
            print(f"records 构建失败：{e}")

        self.records_ready_signal.emit(records)

        os.remove(all_output_txt_file)

        header = ['Index', 'Date', 'Time', 'TimeStamp', 'Count', 'EcuID','Apid', 'Ctid', 'sessionId', 'Type', 'Subtype', 'Mode','Args', 'PayLoad']
        workbook = openpyxl.load_workbook(output_excel)
        worksheet = workbook.active
        worksheet.insert_rows(1)
        for col_idx, col_title in enumerate(header, start=1):
            worksheet.cell(row=1, column=col_idx, value=col_title)
        bold_font = Font(bold=True)
        for col_idx in range(1, len(header) + 1):
            worksheet.cell(row=1, column=col_idx).font = bold_font
        worksheet.freeze_panes = 'A2'
        columns_to_hide = ['C', 'H', 'I', 'J', 'K', 'L', 'M']
        for col in columns_to_hide:
            worksheet.column_dimensions[col].hidden = True
        worksheet.column_dimensions['B'].width = 32
        worksheet.column_dimensions['D'].width = 13

        darkorange = PatternFill(start_color="E26B0A", end_color="00008B", fill_type="solid")
        red = PatternFill(start_color="FFD699", end_color="FF0000", fill_type="solid")
        dark_blue = PatternFill(start_color="6495ED", end_color="00008B", fill_type="solid")
        qianqing = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")

        for row in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=14)
            if "RT1_O" in str(cell.value):
                cell.fill = red
            elif "RT2_O" in str(cell.value):
                cell.fill = qianqing
            elif "RT1_M" in str(cell.value):
                cell.fill = darkorange
            elif "RT2_M" in str(cell.value):
                cell.fill = dark_blue

        workbook.save(output_excel)

        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))
        end_time = time.time()
        print(f"程序结束于:⏰ {time.ctime(end_time)}")
        print(f"程序总运行时间:⏳ {end_time - start_time:.2f}秒")
        self.finished_signal.emit()

# ===================== 简化主界面（删除无用的DLT Viewer配置） =====================
class DltConverter(QMainWindow):
    def __init__(self):
        super().__init__()
        # 解析预览数据
        self.all_records = []
        self.filtered_records = []
        self.current_page = 0
        self.sort_time_asc = True
        self.unix_ts_min = None
        self.unix_ts_max = None
        self._updating_filter_controls = False
        self.initUI()

    def initUI(self):
        self.setWindowTitle("DLT文件转换为xlsx-无依赖版（解析预览）")
        self.setMinimumSize(1100, 950)

        central = QWidget(self)
        self.setCentralWidget(central)

        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # 顶部输入栏
        top_row = QHBoxLayout()
        self.label_dlt_file = QLabel("DLT文件夹:")
        self.entry_dlt_file = QLineEdit()
        self.button_select_dlt_file = QPushButton("选择文件夹")
        self.button_select_dlt_file.clicked.connect(self.select_dlt_file)

        self.button_convert = QPushButton("转♂换")
        self.button_convert.clicked.connect(self.convert)

        top_row.addWidget(self.label_dlt_file, 0)
        top_row.addWidget(self.entry_dlt_file, 1)
        top_row.addWidget(self.button_select_dlt_file, 0)
        top_row.addWidget(self.button_convert, 0)

        # 使用说明
        self.label_instructions = QLabel(
            "<br>使用说明：<br>"
            "1.将要转换的日志文件放入同一个文件夹<br>"
            "2.日志文件可以是.gz压缩包，也可以是.dlt，也可以是.log<br>"
            "3.转换完成后统统会变为.dlt文件<br>"
            "4.完成后所有日志会被合并进同一个xlsx文件<br>"
            , self
        )
        self.label_instructions.setWordWrap(True)
        self.label_instructions.setStyleSheet("color: black; font-family: 'Microsoft YaHei';")

        main_layout.addLayout(top_row)
        main_layout.addWidget(self.label_instructions)

        # 运行进度与日志
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(20)
        main_layout.addWidget(self.progress_bar)

        log_group = QGroupBox("运行日志")
        log_group_layout = QVBoxLayout(log_group)
        self.text_edit = QTextEdit()
        self.text_edit.setReadOnly(True)
        self.text_edit.setStyleSheet(
            "QTextEdit {background-color: #CCE8CF; border: 2px solid #1e90ff; border-radius: 6px; padding: 10px; font-family: '微软雅黑';}"
        )
        log_group_layout.addWidget(self.text_edit)
        log_group.setMinimumHeight(120)
        main_layout.addWidget(log_group)

        sys.stdout = StreamToTextEdit(self.text_edit)

        # ===== 解析预览面板：时间筛选 + 表格展示 =====
        preview_group = QGroupBox("解析预览（仅显示：时间/相对时间/ecuid/application_id/payload）")
        preview_group_layout = QVBoxLayout(preview_group)
        preview_group_layout.setSpacing(8)

        filter_row = QHBoxLayout()
        self.start_time_edit = QDateTimeEdit()
        self.start_time_edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.start_time_edit.setEnabled(False)

        self.end_time_edit = QDateTimeEdit()
        self.end_time_edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.end_time_edit.setEnabled(False)

        self.apply_filter_button = QPushButton("筛选")
        self.apply_filter_button.setEnabled(False)
        self.apply_filter_button.clicked.connect(self.apply_filter)

        self.reset_filter_button = QPushButton("重置")
        self.reset_filter_button.setEnabled(False)
        self.reset_filter_button.clicked.connect(self.reset_filter)

        self.page_size_spinbox = QSpinBox()
        self.page_size_spinbox.setRange(10, 5000)
        self.page_size_spinbox.setValue(500)
        self.page_size_spinbox.setEnabled(False)
        self.page_size_spinbox.valueChanged.connect(self.apply_filter)

        # 用户调整起止时间时自动刷新（用 _updating_filter_controls 防止 on_records_ready 初始化时重复刷新）
        self.start_time_edit.dateTimeChanged.connect(self.apply_filter)
        self.end_time_edit.dateTimeChanged.connect(self.apply_filter)

        filter_row.addWidget(QLabel("开始时间:"))
        filter_row.addWidget(self.start_time_edit)
        filter_row.addWidget(QLabel("结束时间:"))
        filter_row.addWidget(self.end_time_edit)
        filter_row.addSpacing(10)
        filter_row.addWidget(self.apply_filter_button)
        filter_row.addWidget(self.reset_filter_button)
        filter_row.addSpacing(10)
        filter_row.addWidget(QLabel("每页:"))
        filter_row.addWidget(self.page_size_spinbox)

        self.page_info_label = QLabel("显示 0 条")
        filter_row.addSpacing(10)
        filter_row.addWidget(self.page_info_label, 1)

        preview_group_layout.addLayout(filter_row)

        self.records_table = QTableWidget()
        self.records_table.setColumnCount(6)
        self.records_table.setHorizontalHeaderLabels(
            ["时间", "相对时间", "ecuid", "application_id", "payload", "unix_ts"]
        )
        self.records_table.setWordWrap(True)
        self.records_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.records_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.records_table.horizontalHeader().sectionClicked.connect(self.on_table_header_clicked)
        self.records_table.setColumnHidden(5, True)  # unix_ts 仅用于内部筛选/排序
        self.records_table.horizontalHeader().setStretchLastSection(True)

        # 表格行高会随 payload 换行而变大；分页可避免一次渲染过多
        preview_group_layout.addWidget(self.records_table, 1)

        page_row = QHBoxLayout()
        self.page_nav_prev_button = QPushButton("上一页")
        self.page_nav_prev_button.setEnabled(False)
        self.page_nav_prev_button.clicked.connect(self.prev_page)
        self.page_nav_next_button = QPushButton("下一页")
        self.page_nav_next_button.setEnabled(False)
        self.page_nav_next_button.clicked.connect(self.next_page)

        page_row.addWidget(self.page_nav_prev_button, 0)
        page_row.addWidget(self.page_nav_next_button, 0)
        page_row.addStretch(1)

        preview_group_layout.addLayout(page_row)

        main_layout.addWidget(preview_group, 1)

        # 让整体看起来更统一
        self.setStyleSheet(
            """
            QGroupBox { font-family: 'Microsoft YaHei'; font-weight: bold; border: 1px solid #c0c0c0; border-radius: 6px; margin-top: 6px; }
            QPushButton { font-family: 'Microsoft YaHei'; padding: 6px 14px; background-color: #1e90ff; color: white; border-radius: 6px; }
            QPushButton:disabled { background-color: #c0c0c0; color: #666666; }
            QTableWidget { font-family: 'Consolas'; gridline-color: #e0e0e0; }
            QHeaderView::section { background-color: #f2f2f2; font-weight: bold; border: 1px solid #d0d0d0; }
            """
        )

    def select_dlt_file(self):
        file_path = QFileDialog.getExistingDirectory(self, "选择DLT文件夹")
        if file_path:
            self.entry_dlt_file.setText(file_path)

    def convert(self):
        folder_path = self.entry_dlt_file.text()
        if not os.path.isdir(folder_path):
            QMessageBox.critical(self, "错误", "请选择有效的DLT文件夹")
            return
        self.button_convert.setEnabled(False)
        self.worker_thread = MyWorkerThread(folder_path)
        self.worker_thread.progress_signal.connect(self.update_progress)
        self.worker_thread.records_ready_signal.connect(self.on_records_ready)
        self.worker_thread.finished_signal.connect(self.on_conversion_finished)
        self.worker_thread.start()

    def update_progress(self, progress):
        self.progress_bar.setValue(progress)

    def on_records_ready(self, records):
        # records: [{unix_ts, abs_time, rel_time, ecuid, application_id, payload}, ...]
        self._updating_filter_controls = True
        self.all_records = records or []
        self.filtered_records = list(self.all_records)
        self.current_page = 0
        self.sort_time_asc = True

        if not self.all_records:
            self.records_table.setRowCount(0)
            self.page_info_label.setText("显示 0 条")
            self.apply_filter_button.setEnabled(False)
            self.reset_filter_button.setEnabled(False)
            self.page_size_spinbox.setEnabled(False)
            self.page_nav_prev_button.setEnabled(False)
            self.page_nav_next_button.setEnabled(False)
            return

        self.unix_ts_min = min(r.get("unix_ts") for r in self.all_records if r.get("unix_ts") is not None)
        self.unix_ts_max = max(r.get("unix_ts") for r in self.all_records if r.get("unix_ts") is not None)
        if self.unix_ts_min is None or self.unix_ts_max is None:
            self.unix_ts_min, self.unix_ts_max = 0, 0

        # 设置默认筛选区间（绝对时间）
        self.start_time_edit.setDateTime(QDateTime.fromMSecsSinceEpoch(int(float(self.unix_ts_min) * 1000)))
        self.end_time_edit.setDateTime(QDateTime.fromMSecsSinceEpoch(int(float(self.unix_ts_max) * 1000)))

        self.start_time_edit.setEnabled(True)
        self.end_time_edit.setEnabled(True)
        self.apply_filter_button.setEnabled(True)
        self.reset_filter_button.setEnabled(True)
        self.page_size_spinbox.setEnabled(True)

        self.apply_filter()
        self._updating_filter_controls = False

    def reset_filter(self):
        if self.unix_ts_min is None or self.unix_ts_max is None:
            return
        self._updating_filter_controls = True
        self.sort_time_asc = True
        self.start_time_edit.setDateTime(QDateTime.fromMSecsSinceEpoch(int(float(self.unix_ts_min) * 1000)))
        self.end_time_edit.setDateTime(QDateTime.fromMSecsSinceEpoch(int(float(self.unix_ts_max) * 1000)))
        self.current_page = 0
        self.apply_filter()
        self._updating_filter_controls = False

    def apply_filter(self):
        if not self.all_records:
            return
        if self._updating_filter_controls:
            return

        start_ts = self.start_time_edit.dateTime().toMSecsSinceEpoch() / 1000.0
        end_ts = self.end_time_edit.dateTime().toMSecsSinceEpoch() / 1000.0
        if start_ts > end_ts:
            start_ts, end_ts = end_ts, start_ts

        filtered = []
        for r in self.all_records:
            ut = r.get("unix_ts")
            if ut is None:
                continue
            if start_ts <= ut <= end_ts:
                filtered.append(r)

        filtered.sort(key=lambda x: float(x.get("unix_ts", 0)), reverse=not self.sort_time_asc)
        self.filtered_records = filtered
        self.current_page = 0
        self.refresh_table()

    def on_table_header_clicked(self, logicalIndex):
        # 只允许点击“时间”切换升降序
        if logicalIndex != 0:
            return
        self.sort_time_asc = not self.sort_time_asc
        self.current_page = 0
        self.apply_filter()

    def refresh_table(self):
        total = len(self.filtered_records)
        if total == 0:
            self.records_table.setRowCount(0)
            self.page_nav_prev_button.setEnabled(False)
            self.page_nav_next_button.setEnabled(False)
            self.page_info_label.setText("显示 0 条")
            return

        page_size = int(self.page_size_spinbox.value())
        max_page = (total - 1) // page_size
        if self.current_page > max_page:
            self.current_page = max_page
        if self.current_page < 0:
            self.current_page = 0

        start = self.current_page * page_size
        end = min(start + page_size, total)
        page_records = self.filtered_records[start:end]

        self.records_table.setSortingEnabled(False)
        self.records_table.setRowCount(len(page_records))

        for row_idx, r in enumerate(page_records):
            self.records_table.setItem(row_idx, 0, QTableWidgetItem(str(r.get("abs_time", ""))))
            self.records_table.setItem(row_idx, 1, QTableWidgetItem(str(r.get("rel_time", ""))))
            self.records_table.setItem(row_idx, 2, QTableWidgetItem(str(r.get("ecuid", ""))))
            self.records_table.setItem(row_idx, 3, QTableWidgetItem(str(r.get("application_id", ""))))
            self.records_table.setItem(row_idx, 4, QTableWidgetItem(str(r.get("payload", ""))))
            self.records_table.setItem(row_idx, 5, QTableWidgetItem(str(r.get("unix_ts", 0))))

        self.page_nav_prev_button.setEnabled(self.current_page > 0)
        self.page_nav_next_button.setEnabled(self.current_page < max_page)
        self.page_info_label.setText(
            f"显示 {start + 1}-{end} / {total} 条（时间{'升' if self.sort_time_asc else '降'}序）"
        )

    def prev_page(self):
        if self.current_page <= 0:
            return
        self.current_page -= 1
        self.refresh_table()

    def next_page(self):
        page_size = int(self.page_size_spinbox.value())
        total = len(self.filtered_records)
        if total == 0:
            return
        max_page = (total - 1) // page_size
        if self.current_page >= max_page:
            return
        self.current_page += 1
        self.refresh_table()

    def on_conversion_finished(self):
        QMessageBox.information(self, "提示", "转换完成，xlsx文件已保存在dlt文件相同路径的文件夹下。")
        self.button_convert.setEnabled(True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = DltConverter()
    window.show()
    sys.exit(app.exec_())
