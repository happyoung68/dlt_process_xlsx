import sys
import struct
import re
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QPushButton, QMessageBox, QFileDialog, QProgressBar, QTextEdit,
    QMainWindow, QLabel, QLineEdit, QApplication, QWidget, QVBoxLayout,
    QHBoxLayout, QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView,
    QGroupBox, QAbstractItemView, QSpinBox, QDateTimeEdit, QScrollArea,
    QFrame, QGridLayout, QSplitter
)
from PyQt5.QtCore import QSettings, QThread, pyqtSignal, Qt, QDateTime, QTime, QRectF
from PyQt5.QtGui import QColor, QBrush, QFont, QPainter, QPen, QLinearGradient
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
import gzip
import shutil
import time
import io
import tarfile
from typing import List, Tuple, Optional, Dict
import traceback

RT1_signals = [
    ["", "", "", "NMiDPSNMKeepAwakeFlag", "NMEEMKeepAwakeFlag", "NMEEMKeepAwakeFlag_Aux",
     "NMVCUGlobalKeepAwakeFlag", "NMVCUChargingKeepAwakeFlag(新能源,VCU_hcu)"],
    ["DoIPActiveLineStatus", "NMBMSGlobalKeepAwakeFlag", "NMBMSChargingKeepAwakeFlag",
     "NMBMSThermalRunawayKeepAwakeFlag(新能源,BMS)", "-", "NMBMSThermalRunawayKeepAwakeFlag(动力,VCU_slccu)",
     "NMVCUChargingKeepAwakeFlag(动力,VCU_slccu)", "NMVCU2GlobalKeepAwakeFlag"],
    ["NMVHMKeepAwakeFlag", ".", "CSM_RC_KeepAwake", "NMHVACKeepAwakeFlag(新能源,TMS)", ".", ".",
     "BDC_NMBDCKeepAwakeFlag", "NMTMSChargingKeepAwakeFlag"],
    ["NMTMSGlobalKeepAwakeFlag", "NMHMPNCDownloadKeepAwakeFlag", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    ["NMWCCDownloadRequest_RT1", "CAN7_DiagMSG_ID != 0x74F(PNCGlobal)", "CAN7_DiagMSG_ID == 0x74F(RequestOBD)", ".",
     ".", ".", ".", "."],
    ["SPME触发A核网络维持", "SMM触发A核网络维持", "CSM触发A核网络维持", "OTACLT触发A核网络维持", "VCP触发A核网络维持",
     "BDC触发A核网络维持", "VT触发A核网络维持", "."],
    [".", ".", ".", ".", ".", ".", ".", "."]
]

Event_Source_signals = ["PNCGlobal", "PNCCharging", "PNCBSM", "PNCDownload", "PNCEnter", "PNCHazard", "Reserved",
                        "PNCRearlight"]
Acore_signals = ["SPME(PNCGlobal)", "SMM(PNCGlobal)", "CSM(PNCGlobal)", "OTACLT(PNCDownload)", "VCP(PNCGlobal)",
                 "BDC(PNCDownload,PNCGlobal)", "VT(PNCGlobal)", "UA"]

RT2_signals = [
    ["NMATWSKeepAwakeFlag", "NMExteriorLightKeepAwakeFlag", "NMLockingKeepAwakeFlag", "PWL_NMPWLKeepAwakeFlag",
     "SR_NMSunRoofKeepAwakeFlag", "NMPLGKeepAwakeFlag", "NMPowerOperatedDoorKeepAwakeFlag", "MR_NMMirrorKeepAwakeFlag"],
    ["SSW_NMSeatControlKeepAwakeFlag", "NMHVACKeepAwakeFlag", "NMHVACRLSLinAwakeFlag", "NMAQCKeepAwakeFlag",
     "NMTPMSKeepAwakeFlag", "ILC_NMKeepAwakeFlag", "NMESMKeepAwakeFlag", "NMKAEKeepAwakeFlag"],
    ["NMVRMPNCCharingKeepAwakeFlag", "SMM_NMLPDKeepAwakeFlag", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    ["NMWCCDownloadRequest_RT2", ".", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."]
]


class LogParser:
    MAGIC_HEADER = b'DLT\x01'
    MESSAGE_TYPE_MAP = {0x00: "LOG", 0x01: "APP_TRACE", 0x02: "NW_TRACE", 0x03: "CONTROL"}
    LOG_LEVEL_MAP = {0x01: "FATAL", 0x02: "ERROR", 0x03: "WARN", 0x04: "INFO", 0x05: "DEBUG", 0x06: "VERBOSE"}
    DATA_SIZE_MAP = {0x01: 1, 0x02: 2, 0x03: 4, 0x04: 8}
    ENCODING_MAP = {0x00: 'ascii', 0x01: 'utf-8', 0x02: 'hex_ascii', 0x04: 'utf-8'}
    CAL_TIMESTAMP_PATTERN = re.compile(r'\[\#:(\d+\.\d+)s\]')

    @staticmethod
    def parse_storage_header(hex_data: bytes) -> Optional[Dict]:
        if len(hex_data) < 16 or hex_data[:4] != b'DLT\x01':
            return None
        try:
            seconds, microseconds = struct.unpack('<Ii', hex_data[4:12])
            ecuid = hex_data[12:16].decode('ascii', errors='ignore')
            abs_time = datetime.fromtimestamp(seconds + microseconds / 1000000.0)
            return {
                'time': abs_time,
                'time_str': f"{abs_time.strftime('%Y/%m/%d %H:%M:%S.')}{microseconds:06d}",
                'ecuid': ecuid
            }
        except:
            return None

    @staticmethod
    def parse_standard_header(hex_data: bytes) -> Optional[Dict]:
        if len(hex_data) < 16:
            return None
        try:
            header_type_byte = hex_data[0]
            with_ecuid = (header_type_byte & 0x04) != 0
            with_session_id = (header_type_byte & 0x08) != 0
            with_timestamp = (header_type_byte & 0x10) != 0
            msb_first = (header_type_byte & 0x02) != 0
            length = struct.unpack('>H', hex_data[2:4])[0]
            offset = 4
            ecuid = ""
            if with_ecuid and offset + 4 <= len(hex_data):
                ecuid = hex_data[offset:offset + 4].decode('ascii', errors='ignore')
                offset += 4
            session_id = 0
            if with_session_id and offset + 4 <= len(hex_data):
                session_id = struct.unpack('>I', hex_data[offset:offset + 4])[0]
                offset += 4
            timestamp = 0.0
            if with_timestamp and offset + 4 <= len(hex_data):
                timestamp = struct.unpack('>I', hex_data[offset:offset + 4])[0] / 10000.0
            return {
                'msb_first': msb_first,
                'length': length,
                'ecuid': ecuid,
                'session_id': session_id,
                'timestamp': timestamp
            }
        except:
            return None

    @staticmethod
    def parse_extended_header(hex_data: bytes) -> Optional[Dict]:
        if len(hex_data) < 10:
            return None
        try:
            message_info = hex_data[0]
            mode = (message_info & 0x01) != 0
            message_type_bits = (message_info >> 1) & 0x07
            message_type_info_bits = (message_info >> 4) & 0x0F
            message_type = LogParser.MESSAGE_TYPE_MAP.get(message_type_bits, f"RESERVED_{message_type_bits}")
            message_type_info = LogParser.LOG_LEVEL_MAP.get(message_type_info_bits, "") if message_type == "LOG" else ""
            return {
                'mode': mode,
                'message_type': message_type,
                'message_type_info': message_type_info,
                'num_arguments': hex_data[1],
                'application_id': hex_data[2:6].decode('ascii', errors='ignore').replace('\x00', ''),
                'context_id': hex_data[6:10].decode('ascii', errors='ignore').replace('\x00', '')
            }
        except:
            return None

    @staticmethod
    def parse_payload(hex_data: bytes, num_arguments: int, msb_first: bool, mode: bool) -> str:
        if not mode:
            return LogParser._parse_non_verbose_payload(hex_data)
        payload_parts = []
        offset = 0
        total_length = len(hex_data)
        for _arg_idx in range(num_arguments):
            if offset + 4 > total_length:
                break
            type_info = int.from_bytes(hex_data[offset:offset + 4], 'little')
            offset += 4
            data_type_length = type_info & 0x0F
            is_sint = (type_info & 0x20) != 0
            is_usint = (type_info & 0x40) != 0
            is_string = (type_info & 0x200) != 0
            string_code = (type_info >> 15) & 0x07
            result = ""
            if string_code == 0x04:
                if offset + 2 <= total_length:
                    str_len = struct.unpack('>H', hex_data[offset:offset + 2])[0]
                    offset += 2
                    if str_len > 0 and offset + str_len <= total_length:
                        result = hex_data[offset:offset + str_len].decode('utf-8', errors='ignore')
                        offset += str_len
                payload_parts.append(result)
                break
            elif is_string:
                if offset + 2 <= total_length:
                    str_len = struct.unpack('<H', hex_data[offset:offset + 2])[0]
                    offset += 2
                    if str_len > 0 and offset + str_len <= total_length:
                        string_data = hex_data[offset:offset + str_len]
                        offset += str_len
                        encoding = LogParser.ENCODING_MAP.get(string_code, 'ascii')
                        if encoding == 'hex_ascii':
                            result = f"0x{(string_data if msb_first else string_data[::-1]).hex().upper()}"
                        else:
                            result = string_data.decode(encoding, errors='ignore')
            elif is_sint or is_usint:
                data_size = LogParser.DATA_SIZE_MAP.get(data_type_length, 0)
                if data_size > 0 and offset + data_size <= total_length:
                    data_bytes = hex_data[offset:offset + data_size]
                    offset += data_size
                    if is_usint and data_size in (1, 2, 4):
                        try:
                            fmt = f"{'>' if msb_first else '<'}{'BHI'[data_size // 2]}"
                            result = str(struct.unpack(fmt, data_bytes)[0])
                        except:
                            pass
                        if string_code == 0x02:
                            result = f"0x{(data_bytes if msb_first else data_bytes[::-1]).hex().upper()}"
            else:
                if offset < total_length:
                    result = hex_data[offset:].hex(' ', 1)
                offset = total_length
            payload_parts.append(result)
        return " ".join(payload_parts).replace('\x00', '').strip()

    @staticmethod
    def _parse_non_verbose_payload(hex_data: bytes) -> str:
        try:
            if len(hex_data) < 4:
                return ""
            message_id = struct.unpack('>I', hex_data[:4])[0]
            data_bytes = hex_data[4:]
            sep = data_bytes.find(b'DLT\x01')
            if sep != -1:
                data_bytes = data_bytes[:sep]
            if data_bytes:
                return f"[{message_id}] {data_bytes.decode('ascii', errors='ignore')}"
            return f"[{message_id}]"
        except:
            return ""

    @staticmethod
    def split_dlt_messages(file_data: bytes) -> List[bytes]:
        messages = []
        start = 0
        magic = b'DLT\x01'
        file_len = len(file_data)
        while start < file_len:
            next_start = file_data.find(magic, start + 4)
            if next_start == -1:
                messages.append(file_data[start:])
                break
            messages.append(file_data[start:next_start])
            start = next_start
        return messages

    @staticmethod
    def parse_dlt_message(hex_data: bytes, msg_index: int) -> Optional[Dict]:
        if len(hex_data) < 42:
            return None
        try:
            storage_header = LogParser.parse_storage_header(hex_data[:16])
            if not storage_header:
                return None
            standard_header = LogParser.parse_standard_header(hex_data[16:32])
            if not standard_header:
                return None
            payload_length = standard_header['length'] - 16 - 10
            if payload_length < 0:
                return None
            extended_header = LogParser.parse_extended_header(hex_data[32:42])
            if not extended_header:
                return None
            payload_data = hex_data[42:42 + payload_length]
            payload = LogParser.parse_payload(
                payload_data, extended_header['num_arguments'],
                standard_header['msb_first'], extended_header['mode']
            )
            cal_timestamp = standard_header['timestamp']
            if storage_header['ecuid'] in ["CCU0", "CCU1"]:
                match = LogParser.CAL_TIMESTAMP_PATTERN.search(payload)
                if match:
                    cal_timestamp = round(float(match.group(1)), 4)
            return {
                'index': msg_index,
                'time': storage_header['time'],
                'time_str': storage_header['time_str'],
                'timestamp': standard_header['timestamp'],
                'cal_timestamp': cal_timestamp,
                'ecuid': storage_header['ecuid'],
                'application_id': extended_header['application_id'],
                'context_id': extended_header['context_id'],
                'session_id': standard_header['session_id'],
                'message_type': extended_header['message_type'],
                'message_type_info': extended_header['message_type_info'],
                'mode': 'verbose' if extended_header['mode'] else 'non-verbose',
                'payload': payload
            }
        except:
            return None


def parse_dlt_file_fast(dlt_file_path: str) -> List[Dict]:
    if not os.path.exists(dlt_file_path):
        return []
    with open(dlt_file_path, 'rb') as f:
        file_data = f.read()
    messages = LogParser.split_dlt_messages(file_data)
    parsed_results = []
    for idx, msg in enumerate(messages):
        res = LogParser.parse_dlt_message(msg, idx)
        if res:
            parsed_results.append(res)
    return parsed_results


def process_signal_data(parsed_data: List[Dict]) -> List[Dict]:
    for item in parsed_data:
        payload = item.get('payload', '')
        if 'RT1_McoreSignalChange' in payload or 'RT1_NetWUStVar' in payload:
            item['payload'] = process_rt_signal(payload, RT1_signals, True)
        elif 'RT2_McoreSignalChange' in payload or 'RT2_NetWUStVar' in payload:
            item['payload'] = process_rt_signal(payload, RT2_signals, False)
        elif 'EventSource_Current' in payload or 'EvenSource_Current' in payload:
            item['payload'] = process_event_source(payload)
        elif 'AcoreSignalChange' in payload:
            item['payload'] = process_acore_signal(payload)
    return parsed_data


def process_rt_signal(payload: str, signals: List, is_rt1: bool) -> str:
    try:
        if "):" not in payload:
            return payload
        numbers_string = payload.split("):")[1]
        numbers_list = numbers_string.split(",")[:8]
        bin_array = []
        for dec in numbers_list:
            try:
                decimal_value = int(dec.strip(), 16)
                binary_string = bin(decimal_value)[2:].zfill(8)
                bin_array.append(binary_string)
            except:
                continue
        ans = "--"
        if is_rt1 and len(bin_array) > 0:
            car_state = bin_array[0][5] + bin_array[0][6] + bin_array[0][7]
            if car_state == "000":
                ans += "Standby--"
            elif car_state == "001":
                ans += "Comfort--"
            elif car_state == "011":
                ans += "DRV--"
        for i in range(len(bin_array)):
            for j in range(8):
                if bin_array[i][j] == "1" and i < len(signals):
                    sig = signals[i][7 - j]
                    if sig and sig != "." and sig != "-":
                        ans += sig + "--"
        return payload + ans
    except:
        return payload


def process_event_source(payload: str) -> str:
    try:
        if "):" not in payload:
            return payload
        numbers_string = payload.split("):")[1]
        numbers_list = numbers_string.split(",")[:1]
        for dec in numbers_list:
            try:
                val = int(dec.strip(), 16)
                binary = bin(val)[2:].zfill(8)
                ans = "--"
                for j in range(8):
                    if binary[j] == "1":
                        ans += Event_Source_signals[7 - j] + "--"
                return payload + ans
            except:
                continue
        return payload
    except:
        return payload


def process_acore_signal(payload: str) -> str:
    try:
        if "):" not in payload:
            return payload
        numbers_string = payload.split("):")[1]
        numbers_list = numbers_string.split(",")[:1]
        for dec in numbers_list:
            try:
                val = int(dec.strip(), 16)
                binary = bin(val)[2:].zfill(8)
                ans = "--"
                for j in range(8):
                    if binary[j] == "1":
                        ans += Acore_signals[7 - j] + "--"
                return payload + ans
            except:
                continue
        return payload
    except:
        return payload


def extract_hex_bytes_from_payload(payload: str) -> List[int]:
    if "):" not in payload:
        return []
    try:
        hex_part = payload.split("):")[1]
        bytes_str = [b.strip() for b in hex_part.split(",")[:8]]
        return [int(b, 16) for b in bytes_str if b]
    except:
        return []


def get_bit_states_from_bytes(byte_values: List[int]) -> List[int]:
    bits = []
    for val in byte_values[:8]:
        binary = bin(val)[2:].zfill(8)
        for j in range(7, -1, -1):
            bits.append(int(binary[j]))
    return bits


def get_signal_name_from_bit(bit_index: int, signals: List) -> str:
    byte_idx = bit_index // 8
    bit_in_byte = bit_index % 8
    if byte_idx >= len(signals):
        return ""
    sig = signals[byte_idx][bit_in_byte]
    return sig if sig and sig != "." and sig != "-" else ""


def get_vehicle_state_from_byte(byte_val: int) -> str:
    state_val = (byte_val >> 5) & 0x07
    if state_val == 0:
        return "Standby"
    elif state_val == 1:
        return "Comfort"
    elif state_val == 3:
        return "DRV"
    return "Unknown"


def ensure_dlt_extension(filename: str) -> str:
    if '.' not in filename:
        new_filename = filename + '.dlt'
    else:
        base_name, extension = os.path.splitext(filename)
        if extension.lower() != '.dlt':
            new_filename = base_name + '.dlt'
        else:
            return filename
    try:
        os.rename(filename, new_filename)
        return new_filename
    except:
        return filename


def decompress_folder(folder: str):
    temp_folder = os.path.join(folder, 'temp')
    if not os.path.exists(temp_folder):
        os.makedirs(temp_folder)
    for file in os.listdir(folder):
        if file.endswith('.tar.gz'):
            tar_path = os.path.join(folder, file)
            dest_path = os.path.join(temp_folder, file[:-7])
            try:
                import tarfile
                with tarfile.open(tar_path, 'r:gz') as tar:
                    tar.extractall(path=dest_path)
                os.remove(tar_path)
            except:
                pass
    for root, dirs, files in os.walk(folder, topdown=False):
        for file in files:
            file_path = os.path.join(root, file)
            if 'WCC' in file or 'POWM' in file:
                try:
                    shutil.move(file_path, os.path.join(folder, file))
                except:
                    pass
            else:
                try:
                    os.remove(file_path)
                except:
                    pass
        for d in dirs:
            try:
                os.rmdir(os.path.join(root, d))
            except:
                pass
    for file in os.listdir(folder):
        if file.endswith('.gz'):
            gz_path = os.path.join(folder, file)
            out_path = os.path.join(folder, os.path.splitext(file)[0])
            if not os.path.exists(out_path + ".dlt"):
                try:
                    with gzip.open(gz_path, 'rb') as f_in:
                        with open(out_path, 'wb') as f_out:
                            shutil.copyfileobj(f_in, f_out)
                except:
                    pass


class StreamToTextEdit(io.StringIO):
    def __init__(self, text_edit):
        super().__init__()
        self.text_edit = text_edit

    def write(self, text):
        self.text_edit.insertPlainText(text)
        self.text_edit.ensureCursorVisible()


class WorkerThread(QThread):
    progress_signal = pyqtSignal(int, str)
    finished_signal = pyqtSignal(list)
    error_signal = pyqtSignal(str)
    log_signal = pyqtSignal(str)

    def __init__(self, folder_path: str, parent=None):
        super().__init__(parent)
        self.folder_path = folder_path

    def run(self):
        start_time = time.time()
        self.log_signal.emit(f"开始处理: {time.ctime(start_time)}")
        try:
            decompress_folder(self.folder_path)
            files = os.listdir(self.folder_path)
            dlt_files = [f for f in files if f.endswith('.dlt') or f.endswith('.log') or '.' not in f]
            for i in range(len(dlt_files)):
                dlt_files[i] = ensure_dlt_extension(os.path.join(self.folder_path, dlt_files[i]))
            self.progress_signal.emit(10, "解压完成，开始解析DLT文件...")
            all_data = []
            total = len(dlt_files)
            for idx, dlt_file in enumerate(dlt_files):
                self.log_signal.emit(f"解析: {os.path.basename(dlt_file)}")
                data = parse_dlt_file_fast(dlt_file)
                data = process_signal_data(data)
                all_data.extend(data)
                progress = 10 + int((idx + 1) / total * 60) if total > 0 else 70
                self.progress_signal.emit(progress, f"已解析 {idx + 1}/{total} 个文件")
            self.progress_signal.emit(75, "生成Excel...")
            if all_data:
                self.save_to_excel(all_data)
            self.progress_signal.emit(100, "完成")
            self.log_signal.emit(f"完成，耗时: {time.time() - start_time:.2f}秒，共 {len(all_data)} 条")
            self.finished_signal.emit(all_data)
        except Exception as e:
            self.error_signal.emit(f"错误: {str(e)}\n{traceback.format_exc()}")

    def save_to_excel(self, data: List[Dict]) -> str:
        now = datetime.now()
        ts = now.strftime("%Y-%m-%d_%H-%M-%S")
        out = os.path.join(self.folder_path, f"{ts}.xlsx")
        columns_order = ['time_str', 'timestamp', 'ecuid', 'application_id', 'context_id', 
                         'session_id', 'message_type', 'message_type_info', 'payload']
        df_data = []
        for item in data:
            row = {k: item.get(k, '') for k in columns_order}
            df_data.append(row)
        df = pd.DataFrame(df_data, columns=columns_order)
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=True, sheet_name='Logs')
            ws = writer.sheets['Logs']
            headers = ['时间', '相对时间(s)', 'EcuID', 'AppID', 'CtID', 'SessionID', 'Type', 'SubType', 'Payload']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
            ws.freeze_panes = 'A2'
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['I'].width = 100
            ws.column_dimensions['E'].hidden = True
            ws.column_dimensions['F'].hidden = True
            ws.column_dimensions['G'].hidden = True
            ws.column_dimensions['H'].hidden = True
            red = PatternFill(start_color="FFD699", end_color="FF0000", fill_type="solid")
            orange = PatternFill(start_color="E26B0A", end_color="00008B", fill_type="solid")
            blue = PatternFill(start_color="6495ED", end_color="00008B", fill_type="solid")
            cyan = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=9)
                val = str(cell.value) if cell.value else ""
                if "RT1_O" in val:
                    cell.fill = red
                elif "RT2_O" in val:
                    cell.fill = cyan
                elif "RT1_M" in val:
                    cell.fill = orange
                elif "RT2_M" in val:
                    cell.fill = blue
        self.log_signal.emit(f"Excel: {out}")
        return out


class VehicleStateTimelineWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.state_events = []
        self.min_time = None
        self.max_time = None
        self.setMinimumHeight(50)
        self.setMaximumHeight(60)
        
    def set_data(self, state_events: List[Dict], min_time: datetime, max_time: datetime):
        self.state_events = state_events
        self.min_time = min_time
        self.max_time = max_time
        self.update()
        
    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        if not self.state_events or not self.min_time or not self.max_time:
            painter.setFont(QFont("Microsoft YaHei", 8))
            painter.drawText(20, 20, "暂无车辆状态数据")
            return
        
        margin_left = 100
        margin_right = 15
        margin_top = 18
        margin_bottom = 22
        
        width = self.width() - margin_left - margin_right
        total_seconds = (self.max_time - self.min_time).total_seconds()
        
        if total_seconds <= 0:
            return
        
        painter.setPen(QPen(QColor(80, 80, 80)))
        painter.setFont(QFont("Microsoft YaHei", 8, QFont.Bold))
        painter.drawText(5, 14, "车辆状态")
        
        pen = QPen(QColor(180, 180, 180), 1)
        painter.setPen(pen)
        painter.drawLine(margin_left, margin_top, margin_left, self.height() - margin_bottom)
        painter.drawLine(margin_left, self.height() - margin_bottom, 
                        self.width() - margin_right, self.height() - margin_bottom)
        
        for i in range(6):
            x = margin_left + int(width * i / 5)
            painter.drawLine(x, self.height() - margin_bottom, x, self.height() - margin_bottom + 3)
            t = self.min_time + timedelta(seconds=total_seconds * i / 5)
            time_str = t.strftime("%H:%M:%S")
            painter.setFont(QFont("Microsoft YaHei", 6))
            painter.drawText(x - 20, self.height() - margin_bottom + 10, time_str)
        
        state_colors = {
            "Standby": QColor(46, 204, 113),
            "Comfort": QColor(52, 152, 219),
            "DRV": QColor(231, 76, 60),
            "Unknown": QColor(149, 165, 166)
        }
        
        y = margin_top + 2
        bar_height = 14
        
        for event in self.state_events:
            start_t = event['start_time']
            end_t = event['end_time']
            state = event['state']
            
            start_offset = (start_t - self.min_time).total_seconds()
            end_offset = (end_t - self.min_time).total_seconds()
            
            x1 = margin_left + int(width * start_offset / total_seconds)
            x2 = margin_left + int(width * end_offset / total_seconds)
            
            if x2 - x1 < 2:
                x2 = x1 + 2
            
            color = state_colors.get(state, state_colors["Unknown"])
            gradient = QLinearGradient(x1, y, x1, y + bar_height)
            gradient.setColorAt(0, color.lighter(130))
            gradient.setColorAt(1, color)
            
            painter.fillRect(x1, y, x2 - x1, bar_height, QBrush(gradient))
            
            painter.setPen(QPen(QColor(255, 255, 255)))
            painter.drawRect(x1, y, x2 - x1, bar_height)
            
            if x2 - x1 > 30:
                painter.setFont(QFont("Microsoft YaHei", 7, QFont.Bold))
                painter.drawText(x1 + 3, y + 10, state)


class TimelineWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.timeline_data = {}
        self.vehicle_state_events = []
        self.min_time = None
        self.max_time = None
        self.setMinimumHeight(200)
        self.setMouseTracking(True)
        self.hover_x = None
        self.hover_time = None
        self.margin_left = 150
        self.margin_right = 15
        self.margin_top = 25
        self.margin_bottom = 35
        
    def set_data(self, timeline_data: Dict, vehicle_state_events: List[Dict], min_time: datetime, max_time: datetime):
        self.timeline_data = timeline_data
        self.vehicle_state_events = vehicle_state_events
        self.min_time = min_time
        self.max_time = max_time
        if timeline_data:
            vehicle_rows = 1 if vehicle_state_events else 0
            height = max(200, (len(timeline_data) + vehicle_rows) * 28 + 80)
            self.setMinimumHeight(height)
        self.update()
    
    def mouseMoveEvent(self, event):
        if not self.min_time or not self.max_time:
            return
        x = event.x()
        width = self.width() - self.margin_left - self.margin_right
        if self.margin_left <= x <= self.width() - self.margin_right:
            self.hover_x = x
            total_seconds = (self.max_time - self.min_time).total_seconds()
            offset = (x - self.margin_left) / width
            self.hover_time = self.min_time + timedelta(seconds=total_seconds * offset)
            self.update()
        else:
            self.hover_x = None
            self.hover_time = None
            self.update()
    
    def leaveEvent(self, event):
        self.hover_x = None
        self.hover_time = None
        self.update()
        
    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        if not self.timeline_data or not self.min_time or not self.max_time:
            painter.setFont(QFont("Microsoft YaHei", 9))
            painter.drawText(20, 30, "暂无时间轴数据，请先点击诊断")
            return
        
        margin_left = self.margin_left
        margin_right = self.margin_right
        margin_top = self.margin_top
        margin_bottom = self.margin_bottom
        
        width = self.width() - margin_left - margin_right
        total_seconds = (self.max_time - self.min_time).total_seconds()
        
        if total_seconds <= 0:
            return
        
        painter.setPen(QPen(QColor(80, 80, 80)))
        painter.setFont(QFont("Microsoft YaHei", 9))
        painter.drawText(5, 18, "时间轴视图")
        
        pen = QPen(QColor(180, 180, 180), 1)
        painter.setPen(pen)
        painter.drawLine(margin_left, margin_top, margin_left, self.height() - margin_bottom)
        painter.drawLine(margin_left, self.height() - margin_bottom, 
                        self.width() - margin_right, self.height() - margin_bottom)
        
        for i in range(6):
            x = margin_left + int(width * i / 5)
            painter.drawLine(x, self.height() - margin_bottom, x, self.height() - margin_bottom + 4)
            t = self.min_time + timedelta(seconds=total_seconds * i / 5)
            time_str = t.strftime("%H:%M:%S")
            painter.setFont(QFont("Microsoft YaHei", 7))
            painter.drawText(x - 20, self.height() - margin_bottom + 14, time_str)
        
        row_height = 24
        row_idx = 0
        
        if self.vehicle_state_events:
            y = margin_top + row_idx * row_height + row_height // 2
            
            painter.setPen(QPen(QColor(50, 50, 50)))
            painter.setFont(QFont("Microsoft YaHei", 8))
            painter.drawText(5, y + 4, "车辆状态")
            
            state_colors = {
                "Standby": QColor(46, 204, 113),
                "Comfort": QColor(52, 152, 219),
                "DRV": QColor(231, 76, 60),
                "Unknown": QColor(149, 165, 166)
            }
            
            for event in self.vehicle_state_events:
                start_t = event['start_time']
                end_t = event['end_time']
                state = event['state']
                
                start_offset = (start_t - self.min_time).total_seconds()
                end_offset = (end_t - self.min_time).total_seconds()
                
                x1 = margin_left + int(width * start_offset / total_seconds)
                x2 = margin_left + int(width * end_offset / total_seconds)
                
                if x2 - x1 < 2:
                    x2 = x1 + 2
                
                color = state_colors.get(state, state_colors["Unknown"])
                gradient = QLinearGradient(x1, y - 7, x1, y + 7)
                gradient.setColorAt(0, color.lighter(130))
                gradient.setColorAt(1, color)
                
                painter.fillRect(x1, y - 7, x2 - x1, 14, QBrush(gradient))
                
                painter.setPen(QPen(QColor(255, 255, 255)))
                if x2 - x1 > 25:
                    painter.setFont(QFont("Microsoft YaHei", 6, QFont.Bold))
                    painter.drawText(x1 + 2, y + 3, state)
            
            row_idx += 1
            
            painter.setPen(QPen(QColor(200, 200, 200), 1, Qt.DashLine))
            painter.drawLine(margin_left, margin_top + row_idx * row_height, 
                           self.width() - margin_right, margin_top + row_idx * row_height)
            painter.setPen(QPen(QColor(50, 50, 50)))
        
        colors = [
            QColor(231, 76, 60),
            QColor(46, 204, 113),
            QColor(52, 152, 219),
            QColor(155, 89, 182),
            QColor(241, 196, 15),
            QColor(230, 126, 34),
            QColor(26, 188, 156),
            QColor(233, 30, 99),
        ]
        
        for signal_name, events in sorted(self.timeline_data.items()):
            y = margin_top + row_idx * row_height + row_height // 2
            
            painter.setPen(QPen(QColor(50, 50, 50)))
            painter.setFont(QFont("Microsoft YaHei", 8))
            display_name = signal_name if len(signal_name) <= 18 else signal_name[:15] + "..."
            painter.drawText(5, y + 4, display_name)
            
            color = colors[(row_idx - 1) % len(colors)] if self.vehicle_state_events else colors[row_idx % len(colors)]
            
            for event in events:
                start_t = event['activate_time']
                end_t = event['release_time']
                
                start_offset = (start_t - self.min_time).total_seconds()
                end_offset = (end_t - self.min_time).total_seconds()
                
                x1 = margin_left + int(width * start_offset / total_seconds)
                x2 = margin_left + int(width * end_offset / total_seconds)
                
                if x2 - x1 < 2:
                    x2 = x1 + 2
                
                gradient = QLinearGradient(x1, y - 7, x1, y + 7)
                gradient.setColorAt(0, color.lighter(120))
                gradient.setColorAt(1, color)
                
                painter.fillRect(x1, y - 7, x2 - x1, 14, QBrush(gradient))
                
                painter.setPen(QPen(QColor(255, 255, 255), 1))
                duration_min = event['duration_min']
                if x2 - x1 > 25:
                    painter.setFont(QFont("Microsoft YaHei", 6))
                    painter.drawText(x1 + 2, y + 3, f"{duration_min:.0f}分")
            
            row_idx += 1
        
        if self.hover_x is not None and self.hover_time is not None:
            painter.setPen(QPen(QColor(231, 76, 60), 1, Qt.DashLine))
            painter.drawLine(self.hover_x, margin_top, self.hover_x, self.height() - margin_bottom)
            
            time_str = self.hover_time.strftime("%H:%M:%S")
            date_str = self.hover_time.strftime("%Y/%m/%d")
            
            active_sources = []
            for signal_name, events in self.timeline_data.items():
                for event in events:
                    if event['activate_time'] <= self.hover_time <= event['release_time']:
                        active_sources.append(signal_name)
                        break
            
            vehicle_state = None
            for event in self.vehicle_state_events:
                if event['start_time'] <= self.hover_time <= event['end_time']:
                    vehicle_state = event['state']
                    break
            
            painter.setPen(QPen(QColor(231, 76, 60)))
            painter.setFont(QFont("Microsoft YaHei", 8, QFont.Bold))
            
            max_src_len = max((len(src) for src in active_sources), default=0)
            text_width = max(180, max_src_len * 8 + 20)
            line_height = 14
            base_height = 38
            num_sources_to_show = min(len(active_sources), 5)
            extra_height = 0
            if vehicle_state:
                extra_height += line_height
            if active_sources:
                extra_height += line_height + num_sources_to_show * (line_height - 2)
                if len(active_sources) > 5:
                    extra_height += line_height - 2
            text_height = base_height + extra_height
            
            text_x = self.hover_x + 5
            text_y = margin_top + 5
            
            if text_x + text_width > self.width() - margin_right:
                text_x = self.hover_x - text_width - 5
            
            painter.fillRect(text_x - 3, text_y - 3, text_width, text_height, QBrush(QColor(255, 255, 255, 240)))
            painter.setPen(QPen(QColor(231, 76, 60), 1))
            painter.drawRect(text_x - 3, text_y - 3, text_width, text_height)
            
            painter.setPen(QPen(QColor(50, 50, 50)))
            painter.setFont(QFont("Microsoft YaHei", 8))
            painter.drawText(text_x, text_y + 12, f"{date_str}")
            painter.setPen(QPen(QColor(231, 76, 60)))
            painter.setFont(QFont("Microsoft YaHei", 9, QFont.Bold))
            painter.drawText(text_x, text_y + 26, f"{time_str}")
            
            line_y = text_y + 42
            if vehicle_state:
                state_colors = {
                    "Standby": QColor(46, 204, 113),
                    "Comfort": QColor(52, 152, 219),
                    "DRV": QColor(231, 76, 60),
                    "Unknown": QColor(149, 165, 166)
                }
                painter.setPen(QPen(state_colors.get(vehicle_state, QColor(100, 100, 100))))
                painter.setFont(QFont("Microsoft YaHei", 7))
                painter.drawText(text_x, line_y, f"状态: {vehicle_state}")
                line_y += line_height
            
            if active_sources:
                painter.setPen(QPen(QColor(231, 76, 60)))
                painter.setFont(QFont("Microsoft YaHei", 7, QFont.Bold))
                painter.drawText(text_x, line_y, f"激活维持源 ({len(active_sources)}):")
                line_y += line_height
                painter.setFont(QFont("Microsoft YaHei", 6))
                painter.setPen(QPen(QColor(80, 80, 80)))
                for src in active_sources[:5]:
                    painter.drawText(text_x + 5, line_y, src)
                    line_y += line_height - 2
                if len(active_sources) > 5:
                    painter.drawText(text_x + 5, line_y, f"...等{len(active_sources)}个")


class NoSleepDiagnosisWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.all_data = []
        self.unix_ts_min = None
        self.unix_ts_max = None
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)

        filter_group = QGroupBox("诊断时间段")
        fl = QHBoxLayout(filter_group)
        fl.addWidget(QLabel("开始:"))
        self.start_input = QDateTimeEdit()
        self.start_input.setDisplayFormat("yyyy/MM/dd HH:mm:ss")
        self.start_input.setCalendarPopup(True)
        self.start_input.setFixedWidth(160)
        now = QDateTime.currentDateTime()
        self.start_input.setDateTime(QDateTime(now.date(), QTime(0, 0, 0)))
        fl.addWidget(self.start_input)
        fl.addWidget(QLabel("结束:"))
        self.end_input = QDateTimeEdit()
        self.end_input.setDisplayFormat("yyyy/MM/dd HH:mm:ss")
        self.end_input.setCalendarPopup(True)
        self.end_input.setFixedWidth(160)
        self.end_input.setDateTime(QDateTime(now.date(), QTime(23, 59, 59)))
        fl.addWidget(self.end_input)
        fl.addWidget(QLabel("阈值(分):"))
        self.threshold_spin = QSpinBox()
        self.threshold_spin.setRange(1, 9999)
        self.threshold_spin.setValue(30)
        fl.addWidget(self.threshold_spin)
        self.diag_btn = QPushButton("诊断")
        self.diag_btn.clicked.connect(self.run_diagnosis)
        fl.addWidget(self.diag_btn)
        fl.addStretch()
        layout.addWidget(filter_group)

        timeline_group = QGroupBox("时间轴视图 (车辆状态 + 超时维持源)")
        timeline_layout = QVBoxLayout(timeline_group)
        self.timeline_widget = TimelineWidget()
        timeline_layout.addWidget(self.timeline_widget)
        layout.addWidget(timeline_group)

        result_group = QGroupBox("诊断结果 (每次超阈值的维持事件)")
        rl = QVBoxLayout(result_group)
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(6)
        self.result_table.setHorizontalHeaderLabels(['维持源', '类型', '维持时长(分)', '激活时间', '释放时间', '状态'])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.result_table.horizontalHeader().setStretchLastSection(True)
        self.result_table.setColumnWidth(0, 250)
        self.result_table.setColumnWidth(1, 60)
        self.result_table.setColumnWidth(2, 100)
        self.result_table.setColumnWidth(3, 160)
        self.result_table.setColumnWidth(4, 160)
        self.result_table.setColumnWidth(5, 80)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setStyleSheet("QHeaderView::section { background-color: #e74c3c; color: white; padding: 5px; font-weight: bold; }")
        rl.addWidget(self.result_table)
        layout.addWidget(result_group)

    def set_data(self, data: List[Dict]):
        self.all_data = data
        if data:
            ts_list = [d.get('unix_ts') for d in data if d.get('unix_ts')]
            if ts_list:
                self.unix_ts_min = min(ts_list)
                self.unix_ts_max = max(ts_list)
                self.start_input.setDateTime(QDateTime.fromMSecsSinceEpoch(int(self.unix_ts_min * 1000)))
                self.end_input.setDateTime(QDateTime.fromMSecsSinceEpoch(int(self.unix_ts_max * 1000)))

    def run_diagnosis(self):
        start_dt = self.start_input.dateTime().toPyDateTime()
        end_dt = self.end_input.dateTime().toPyDateTime()
        threshold = self.threshold_spin.value()

        rt1_logs = self.extract_signal_logs(start_dt, end_dt, "RT1")
        rt2_logs = self.extract_signal_logs(start_dt, end_dt, "RT2")

        rt1_results = self.calculate_duration_with_events(rt1_logs, start_dt, end_dt, RT1_signals, "RT1", threshold)
        rt2_results = self.calculate_duration_with_events(rt2_logs, start_dt, end_dt, RT2_signals, "RT2", threshold)

        all_results = rt1_results + rt2_results
        all_results.sort(key=lambda x: (x['signal'], x['activate_time']))

        self.update_result_table(all_results)
        self.update_timeline(all_results, rt1_logs, start_dt, end_dt)

    def update_timeline(self, results: List[Dict], rt1_logs: List[Dict], min_time: datetime, max_time: datetime):
        timeline_data = {}
        for r in results:
            signal = r['signal']
            if signal not in timeline_data:
                timeline_data[signal] = []
            timeline_data[signal].append(r)
        
        vehicle_state_events = self.calculate_vehicle_state_events(rt1_logs, min_time, max_time)
        
        self.timeline_widget.set_data(timeline_data, vehicle_state_events, min_time, max_time)

    def calculate_vehicle_state_events(self, rt1_logs: List[Dict], min_time: datetime, max_time: datetime) -> List[Dict]:
        state_events = []
        current_state = None
        state_start = None
        
        for log in rt1_logs:
            if not log['bits'] or len(log['bits']) < 8:
                continue
            
            state_val = (log['bits'][2] << 2) | (log['bits'][1] << 1) | log['bits'][0]
            
            if state_val == 0:
                state = "Standby"
            elif state_val == 1:
                state = "Comfort"
            elif state_val == 3:
                state = "DRV"
            else:
                state = "Unknown"
            
            if state != current_state:
                if current_state is not None and state_start is not None:
                    state_events.append({
                        'state': current_state,
                        'start_time': state_start,
                        'end_time': log['time']
                    })
                current_state = state
                state_start = log['time']
        
        if current_state is not None and state_start is not None:
            state_events.append({
                'state': current_state,
                'start_time': state_start,
                'end_time': max_time
            })
        
        return state_events

    def extract_signal_logs(self, start_dt: datetime, end_dt: datetime, sig_type: str) -> List[Dict]:
        logs = []
        kw1 = f"{sig_type}_McoreSignalChange"
        kw2 = f"{sig_type}_NetWUStVar"
        for item in self.all_data:
            payload = item.get('payload', '')
            if kw1 not in payload and kw2 not in payload:
                continue
            try:
                item_dt = datetime.strptime(item['time_str'][:19], "%Y/%m/%d %H:%M:%S")
            except:
                continue
            if item_dt < start_dt or item_dt > end_dt:
                continue
            bytes_val = extract_hex_bytes_from_payload(payload)
            if not bytes_val:
                continue
            bits = get_bit_states_from_bytes(bytes_val)
            logs.append({
                'time': item_dt,
                'time_str': item['time_str'],
                'bits': bits,
                'type': sig_type
            })
        logs.sort(key=lambda x: x['time'])
        return logs

    def calculate_duration_with_events(self, logs: List[Dict], start_dt: datetime, end_dt: datetime, 
                                        signals: List, sig_type: str, threshold_min: float) -> List[Dict]:
        threshold_sec = threshold_min * 60
        active_start = [None] * 64
        results = []

        for log in logs:
            t = log['time']
            cur_bits = log['bits']
            for i in range(64):
                cur_state = cur_bits[i] if i < len(cur_bits) else 0
                prev_state = 1 if active_start[i] is not None else 0
                
                if cur_state == 1 and prev_state == 0:
                    active_start[i] = t
                elif cur_state == 0 and prev_state == 1:
                    if active_start[i] is not None:
                        duration = (t - active_start[i]).total_seconds()
                        if duration >= threshold_sec:
                            sig_name = get_signal_name_from_bit(i, signals)
                            if sig_name and "Standby" not in sig_name:
                                results.append({
                                    'signal': sig_name,
                                    'type': sig_type,
                                    'duration_sec': duration,
                                    'duration_min': duration / 60.0,
                                    'activate_time': active_start[i],
                                    'release_time': t,
                                    'is_active': False
                                })
                    active_start[i] = None

        for i in range(64):
            if active_start[i] is not None:
                duration = (end_dt - active_start[i]).total_seconds()
                if duration >= threshold_sec:
                    sig_name = get_signal_name_from_bit(i, signals)
                    if sig_name and "Standby" not in sig_name:
                        results.append({
                            'signal': sig_name,
                            'type': sig_type,
                            'duration_sec': duration,
                            'duration_min': duration / 60.0,
                            'activate_time': active_start[i],
                            'release_time': end_dt,
                            'is_active': True
                        })

        return results

    def update_result_table(self, results: List[Dict]):
        self.result_table.setRowCount(len(results))
        for row, r in enumerate(results):
            self.result_table.setItem(row, 0, QTableWidgetItem(r['signal']))
            self.result_table.setItem(row, 1, QTableWidgetItem(r['type']))
            dur_item = QTableWidgetItem(f"{r['duration_min']:.1f}")
            dur_item.setBackground(QBrush(QColor(231, 76, 60)))
            dur_item.setForeground(QBrush(QColor(255, 255, 255)))
            self.result_table.setItem(row, 2, dur_item)
            activate_str = r['activate_time'].strftime("%Y/%m/%d %H:%M:%S") if r['activate_time'] else ""
            release_str = r['release_time'].strftime("%Y/%m/%d %H:%M:%S") if r['release_time'] else ""
            self.result_table.setItem(row, 3, QTableWidgetItem(activate_str))
            self.result_table.setItem(row, 4, QTableWidgetItem(release_str))
            status = "维持中" if r.get('is_active', False) else "已释放"
            status_item = QTableWidgetItem(status)
            if r.get('is_active', False):
                status_item.setBackground(QBrush(QColor(255, 193, 7)))
            else:
                status_item.setBackground(QBrush(QColor(40, 167, 69)))
                status_item.setForeground(QBrush(QColor(255, 255, 255)))
            self.result_table.setItem(row, 5, status_item)


class LogViewerWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.all_data = []
        self.filtered_data = []
        self.sort_order = Qt.AscendingOrder
        self.unix_ts_min = None
        self.unix_ts_max = None
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)

        filter_group = QGroupBox("时间段筛选")
        fl = QHBoxLayout(filter_group)
        fl.addWidget(QLabel("开始:"))
        self.start_input = QDateTimeEdit()
        self.start_input.setDisplayFormat("yyyy/MM/dd HH:mm:ss")
        self.start_input.setCalendarPopup(True)
        self.start_input.setFixedWidth(160)
        now = QDateTime.currentDateTime()
        self.start_input.setDateTime(QDateTime(now.date(), QTime(0, 0, 0)))
        fl.addWidget(self.start_input)
        fl.addWidget(QLabel("结束:"))
        self.end_input = QDateTimeEdit()
        self.end_input.setDisplayFormat("yyyy/MM/dd HH:mm:ss")
        self.end_input.setCalendarPopup(True)
        self.end_input.setFixedWidth(160)
        self.end_input.setDateTime(QDateTime(now.date(), QTime(23, 59, 59)))
        fl.addWidget(self.end_input)
        self.filter_btn = QPushButton("筛选")
        self.filter_btn.clicked.connect(self.apply_filter)
        fl.addWidget(self.filter_btn)
        self.reset_btn = QPushButton("重置")
        self.reset_btn.clicked.connect(self.reset_filter)
        fl.addWidget(self.reset_btn)
        self.count_label = QLabel("共 0 条")
        fl.addWidget(self.count_label)
        fl.addStretch()
        layout.addWidget(filter_group)

        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(['时间', '相对时间(s)', 'EcuID', 'AppID', 'Payload'])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setColumnWidth(0, 200)
        self.table.setColumnWidth(1, 100)
        self.table.setColumnWidth(2, 80)
        self.table.setColumnWidth(3, 80)
        self.table.setSortingEnabled(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setDefaultSectionSize(24)
        self.table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        self.table.setStyleSheet("""
            QHeaderView::section { background-color: #3498db; color: white; padding: 4px; font-weight: bold; }
            QTableWidget::item { padding: 2px; }
        """)
        layout.addWidget(self.table)

    def set_data(self, data: List[Dict]):
        self.all_data = data
        self.filtered_data = data.copy()
        if data:
            ts_list = [d.get('unix_ts') for d in data if d.get('unix_ts')]
            if ts_list:
                self.unix_ts_min = min(ts_list)
                self.unix_ts_max = max(ts_list)
                self.start_input.setDateTime(QDateTime.fromMSecsSinceEpoch(int(self.unix_ts_min * 1000)))
                self.end_input.setDateTime(QDateTime.fromMSecsSinceEpoch(int(self.unix_ts_max * 1000)))
        self.update_table()

    def update_table(self):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(self.filtered_data))
        for row, item in enumerate(self.filtered_data):
            time_item = QTableWidgetItem(item.get('time_str', ''))
            time_item.setData(Qt.UserRole, item.get('time', datetime.min))
            ts = item.get('timestamp', 0)
            rel_item = QTableWidgetItem(f"{ts:.4f}")
            rel_item.setData(Qt.UserRole, ts)
            ecuid_item = QTableWidgetItem(item.get('ecuid', ''))
            appid_item = QTableWidgetItem(item.get('application_id', ''))
            payload_item = QTableWidgetItem(item.get('payload', ''))
            payload = item.get('payload', '')
            if 'RT1_O' in payload:
                payload_item.setBackground(QBrush(QColor(255, 214, 153)))
            elif 'RT2_O' in payload:
                payload_item.setBackground(QBrush(QColor(224, 255, 255)))
            elif 'RT1_M' in payload:
                payload_item.setBackground(QBrush(QColor(226, 107, 10)))
                payload_item.setForeground(QBrush(QColor(255, 255, 255)))
            elif 'RT2_M' in payload:
                payload_item.setBackground(QBrush(QColor(100, 149, 237)))
            self.table.setItem(row, 0, time_item)
            self.table.setItem(row, 1, rel_item)
            self.table.setItem(row, 2, ecuid_item)
            self.table.setItem(row, 3, appid_item)
            self.table.setItem(row, 4, payload_item)
        self.table.setSortingEnabled(True)
        self.table.sortByColumn(0, Qt.AscendingOrder)
        self.count_label.setText(f"共 {len(self.filtered_data)} 条")

    def apply_filter(self):
        start_dt = self.start_input.dateTime().toPyDateTime()
        end_dt = self.end_input.dateTime().toPyDateTime()
        filtered = []
        for item in self.all_data:
            try:
                item_dt = datetime.strptime(item['time_str'][:19], "%Y/%m/%d %H:%M:%S")
            except:
                continue
            if start_dt <= item_dt <= end_dt:
                filtered.append(item)
        self.filtered_data = filtered
        self.update_table()

    def reset_filter(self):
        if self.unix_ts_min and self.unix_ts_max:
            self.start_input.setDateTime(QDateTime.fromMSecsSinceEpoch(int(self.unix_ts_min * 1000)))
            self.end_input.setDateTime(QDateTime.fromMSecsSinceEpoch(int(self.unix_ts_max * 1000)))
        self.filtered_data = self.all_data.copy()
        self.update_table()

    def on_header_clicked(self, idx):
        self.table.sortItems(idx, Qt.AscendingOrder if self.sort_order == Qt.DescendingOrder else Qt.DescendingOrder)
        self.sort_order = Qt.DescendingOrder if self.sort_order == Qt.AscendingOrder else Qt.AscendingOrder


class HelpWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 30, 40, 30)
        
        title = QLabel("使用说明")
        title.setStyleSheet("font-size: 28px; font-weight: bold; color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(title)
        
        content = QLabel("""
<span style='font-size: 16px; line-height: 2.0;'>

<b>第一步：准备文件</b><br>
把日志文件放到一个文件夹里，支持 .dlt、.log、.gz、.tar.gz 格式<br><br>

<b>第二步：转换文件</b><br>
点击「选择」按钮选择文件夹，然后点击「转换」按钮，等待完成即可<br>
Excel文件会自动生成，显示时间、相对时间、EcuID、AppID、Payload 这些列<br><br>

<b>第三步：查看日志</b><br>
在「日志查看」标签页可以按时间筛选日志，点击表头可以排序<br>
橙色=RT1_M，蓝色=RT2_M，浅红=RT1_O，青色=RT2_O<br><br>

<b>第四步：不休眠诊断</b><br>
这个功能帮你找出哪些信号让车一直醒着不睡觉<br>
设置时间段和阈值（比如30分钟），点击「诊断」<br>
每次超阈值的事件都会单独列出来，时间轴会直观显示各维持源的激活时段<br>
车辆状态时间轴会显示 Standby、Comfort、DRV 状态变化<br><br>

<b>常见问题</b><br>
Q: 诊断结果为空？ A: 检查时间段、调小阈值、确认日志里有RT1/RT2信号<br>
Q: 怎么看隐藏的列？ A: Excel里选中列标题，右键选"取消隐藏"<br>

</span>
""")
        content.setWordWrap(True)
        content.setTextFormat(Qt.RichText)
        layout.addWidget(content)
        layout.addStretch()


class DltConverterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.settings = QSettings("ZhaixiyangTools", "DLT2XLSX_Diag")
        self.worker_thread = None
        self.parsed_data = []
        self.initUI()
        self.load_settings()

    def initUI(self):
        self.setWindowTitle("DLT日志分析工具 - 不休眠诊断版")
        self.setGeometry(100, 100, 1300, 900)
        self.setStyleSheet("""
            QMainWindow { background-color: #f0f0f0; }
            QGroupBox { font-weight: bold; border: 2px solid #3498db; border-radius: 5px; margin-top: 10px; padding-top: 10px; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; color: #2c3e50; }
            QPushButton { background-color: #3498db; color: white; border: none; padding: 8px 16px; border-radius: 4px; font-weight: bold; }
            QPushButton:hover { background-color: #2980b9; }
            QPushButton:disabled { background-color: #bdc3c7; }
            QLineEdit { padding: 6px; border: 2px solid #bdc3c7; border-radius: 4px; }
            QLineEdit:focus { border-color: #3498db; }
            QProgressBar { border: 2px solid #3498db; border-radius: 5px; text-align: center; background-color: #ecf0f1; }
            QProgressBar::chunk { background-color: #3498db; border-radius: 3px; }
        """)
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(10)

        input_group = QGroupBox("文件选择")
        il = QHBoxLayout(input_group)
        il.addWidget(QLabel("DLT文件夹:"))
        self.folder_input = QLineEdit()
        self.folder_input.setPlaceholderText("选择DLT文件夹...")
        il.addWidget(self.folder_input)
        self.select_btn = QPushButton("选择")
        self.select_btn.clicked.connect(self.select_folder)
        il.addWidget(self.select_btn)
        self.convert_btn = QPushButton("转换")
        self.convert_btn.clicked.connect(self.start_convert)
        il.addWidget(self.convert_btn)
        main_layout.addWidget(input_group)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        main_layout.addWidget(self.progress_bar)

        self.status_label = QLabel("就绪")
        self.status_label.setStyleSheet("color: #2c3e50; font-weight: bold;")
        main_layout.addWidget(self.status_label)

        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane { border: 2px solid #3498db; border-radius: 5px; background-color: white; }
            QTabBar::tab { background-color: #ecf0f1; color: #2c3e50; padding: 8px 20px; margin-right: 2px; border-top-left-radius: 4px; border-top-right-radius: 4px; }
            QTabBar::tab:selected { background-color: #3498db; color: white; }
            QTabBar::tab:hover:!selected { background-color: #bdc3c7; }
        """)
        self.log_viewer = LogViewerWidget()
        self.tab_widget.addTab(self.log_viewer, "日志查看")
        self.no_sleep_diag = NoSleepDiagnosisWidget()
        self.tab_widget.addTab(self.no_sleep_diag, "不休眠诊断")
        log_widget = QWidget()
        log_layout = QVBoxLayout(log_widget)
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        self.log_output.setStyleSheet("QTextEdit { background-color: #1e1e1e; color: #d4d4d4; font-family: 'Consolas'; font-size: 12px; border: 1px solid #3498db; border-radius: 5px; }")
        log_layout.addWidget(self.log_output)
        self.tab_widget.addTab(log_widget, "运行日志")
        self.help_widget = HelpWidget()
        self.tab_widget.addTab(self.help_widget, "使用说明")
        main_layout.addWidget(self.tab_widget)

        sys.stdout = StreamToTextEdit(self.log_output)

    def load_settings(self):
        folder = self.settings.value("DLTFolder", "")
        self.folder_input.setText(folder)

    def save_settings(self):
        self.settings.setValue("DLTFolder", self.folder_input.text())

    def closeEvent(self, event):
        self.save_settings()
        event.accept()

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择DLT文件夹")
        if folder:
            self.folder_input.setText(folder)

    def start_convert(self):
        folder = self.folder_input.text()
        if not os.path.isdir(folder):
            QMessageBox.critical(self, "错误", "请选择有效文件夹")
            return
        self.convert_btn.setEnabled(False)
        self.select_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.parsed_data = []
        self.worker_thread = WorkerThread(folder)
        self.worker_thread.progress_signal.connect(self.update_progress)
        self.worker_thread.finished_signal.connect(self.on_finished)
        self.worker_thread.error_signal.connect(self.on_error)
        self.worker_thread.log_signal.connect(lambda m: self.log_output.append(m))
        self.worker_thread.start()

    def update_progress(self, val: int, msg: str):
        self.progress_bar.setValue(val)
        self.status_label.setText(msg)

    def on_finished(self, data: List[Dict]):
        self.parsed_data = data
        self.log_viewer.set_data(data)
        self.no_sleep_diag.set_data(data)
        self.convert_btn.setEnabled(True)
        self.select_btn.setEnabled(True)
        self.tab_widget.setCurrentIndex(0)
        QMessageBox.information(self, "完成", f"转换完成！共 {len(data)} 条日志")

    def on_error(self, msg: str):
        self.convert_btn.setEnabled(True)
        self.select_btn.setEnabled(True)
        QMessageBox.warning(self, "错误", msg)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = DltConverterApp()
    window.show()
    sys.exit(app.exec_())
