import base64
import sys
from PyQt5.QtWidgets import (QPushButton, QMessageBox, QFileDialog, QProgressBar, 
                             QTextEdit, QTableWidget, QTableWidgetItem, QHeaderView,
                             QTabWidget, QWidget, QVBoxLayout, QHBoxLayout, QDateTimeEdit,
                             QLabel, QComboBox, QSplitter, QAbstractItemView)
from PyQt5.QtCore import QSettings, QThread, pyqtSignal, Qt, QDateTime
from PyQt5.QtWidgets import QMainWindow, QLineEdit, QApplication
from PyQt5.QtGui import QFont, QColor, QBrush
import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
import gzip
import shutil
import time
import io
import tarfile
import struct
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import deque

class LogParser:
    """DLT日志解析器（适配non-verbose模式，无第三方依赖）"""
    MAGIC_HEADER = b'DLT\x01'
    STORAGE_HEADER_LEN = 16
    STANDARD_HEADER_LEN = 16
    EXTENDED_HEADER_LEN = 10
    MIN_MSG_LENGTH = 42

    MESSAGE_TYPE_MAP = {
        0x00: "LOG",
        0x01: "APP_TRACE",
        0x02: "NW_TRACE",
        0x03: "CONTROL"
    }
    LOG_LEVEL_MAP = {
        0x01: "FATAL",
        0x02: "ERROR",
        0x03: "WARN",
        0x04: "INFO",
        0x05: "DEBUG",
        0x06: "VERBOSE"
    }
    DATA_SIZE_MAP = {0x01: 1, 0x02: 2, 0x03: 4, 0x04: 8}
    ENCODING_MAP = {0x00: 'ascii', 0x01: 'utf-8', 0x02: 'hex_ascii', 0x04: 'utf-8'}
    
    STORAGE_HEADER_FMT = '<Ii'
    LENGTH_FMT = '>H'
    SESSION_ID_FMT = '>I'
    TIMESTAMP_FMT = '>I'
    STR_LEN_FMT = '<H'
    NON_VERBOSE_MSG_ID_LEN = 4
    DLT_SEPARATOR = b'DLT\x01'
    CAL_TIMESTAMP_PATTERN = re.compile(r'\[\#:(\d+\.\d+)s\]')

    @classmethod
    def parse_storage_header(cls, hex_data):
        if len(hex_data) < cls.STORAGE_HEADER_LEN:
            raise ValueError("StorageHeader长度不足16字节")
        if hex_data[:4] != cls.MAGIC_HEADER:
            raise ValueError("无效的DLT文件格式")

        seconds, microseconds = struct.unpack(cls.STORAGE_HEADER_FMT, hex_data[4:12])
        ecuid = hex_data[12:16].decode('ascii', errors='ignore')
        abs_time = datetime.fromtimestamp(seconds + microseconds / 1000000.0)
        time_str = f"{abs_time.strftime('%Y/%m/%d %H:%M:%S.')}{microseconds:06d}"

        return {'magic': cls.MAGIC_HEADER.hex(), 'seconds': seconds, 'microseconds': microseconds, 'ecuid': ecuid, 'time': time_str}

    @classmethod
    def parse_standard_header(cls, hex_data):
        if len(hex_data) < cls.STANDARD_HEADER_LEN:
            raise ValueError("StandardHeader长度不足16字节")

        header_type_byte = hex_data[0]
        use_extended_header = (header_type_byte & 0x01) != 0
        msb_first = (header_type_byte & 0x02) != 0
        with_ecuid = (header_type_byte & 0x04) != 0
        with_session_id = (header_type_byte & 0x08) != 0
        with_timestamp = (header_type_byte & 0x10) != 0
        version_number = (header_type_byte >> 5) & 0x07

        message_counter = hex_data[1]
        length = struct.unpack(cls.LENGTH_FMT, hex_data[2:4])[0]

        offset = 4
        ecuid = ""
        if with_ecuid and offset + 4 <= len(hex_data):
            ecuid = hex_data[offset:offset+4].decode('ascii', errors='ignore')
            offset += 4

        session_id = 0
        if with_session_id and offset + 4 <= len(hex_data):
            session_id = struct.unpack(cls.SESSION_ID_FMT, hex_data[offset:offset+4])[0]
            offset += 4

        timestamp = 0.0
        if with_timestamp and offset + 4 <= len(hex_data):
            timestamp = struct.unpack(cls.TIMESTAMP_FMT, hex_data[offset:offset+4])[0] / 10000.0

        return {'header_type': header_type_byte, 'use_extended_header': use_extended_header, 'msb_first': msb_first,
                'with_ecuid': with_ecuid, 'with_session_id': with_session_id, 'with_timestamp': with_timestamp,
                'version_number': version_number, 'message_counter': message_counter, 'length': length,
                'ecuid': ecuid, 'session_id': session_id, 'timestamp': timestamp}

    @classmethod
    def parse_extended_header(cls, hex_data):
        if len(hex_data) < cls.EXTENDED_HEADER_LEN:
            raise ValueError("ExtendedHeader长度不足10字节")

        message_info = hex_data[0]
        mode = (message_info & 0x01) != 0
        message_type_bits = (message_info >> 1) & 0x07
        message_type_info_bits = (message_info >> 4) & 0x0F

        message_type = cls.MESSAGE_TYPE_MAP.get(message_type_bits, f"RESERVED_{message_type_bits}")
        message_type_info = cls.LOG_LEVEL_MAP.get(message_type_info_bits, "") if message_type == "LOG" else ""

        num_arguments = hex_data[1]
        application_id = hex_data[2:6].decode('ascii', errors='ignore').replace('\x00', '')
        context_id = hex_data[6:10].decode('ascii', errors='ignore').replace('\x00', '')

        return {'message_info': message_info, 'mode': mode, 'message_type': message_type,
                'message_type_info': message_type_info, 'num_arguments': num_arguments,
                'application_id': application_id, 'context_id': context_id}

    @classmethod
    def parse_payload(cls, hex_data, num_arguments, msb_first, mode):
        if not mode:
            return cls._parse_non_verbose_payload(hex_data)

        payload_parts = []
        payload_p_parts = []
        offset = 0
        total_length = len(hex_data)
        data_size_map = cls.DATA_SIZE_MAP
        encoding_map = cls.ENCODING_MAP
        str_len_fmt = cls.STR_LEN_FMT
        max_offset = total_length - 4

        for _arg_idx in range(num_arguments):
            if offset > max_offset:
                break

            type_info = int.from_bytes(hex_data[offset:offset+4], 'little')
            offset += 4

            data_type_length = type_info & 0x0F
            is_bool = (type_info & 0x10) != 0
            is_sint = (type_info & 0x20) != 0
            is_usint = (type_info & 0x40) != 0
            is_float = (type_info & 0x80) != 0
            is_string = (type_info & 0x200) != 0
            is_raw = (type_info & 0x400) != 0
            string_code = (type_info >> 15) & 0x07

            result = ""
            group_part = ""
            is_special_string = False

            if string_code == 0x04:
                is_string = True
                is_special_string = True
                num_arguments = 1

            if is_string:
                if is_special_string:
                    if offset + 2 > total_length:
                        result = ""
                    else:
                        str_len = struct.unpack('>H', hex_data[offset:offset+2])[0]
                        offset += 2
                        if str_len <= 0 or offset + str_len > total_length:
                            result = ""
                        else:
                            string_data = hex_data[offset:offset+str_len]
                            offset += str_len
                            encoding = encoding_map.get(string_code, 'ascii')
                            result = string_data.decode(encoding, errors='ignore')
                            group_part = result
                    payload_parts.append(result)
                    payload_p_parts.append(group_part)
                    break
                else:
                    if offset + 2 > total_length:
                        break
                    str_len = struct.unpack(str_len_fmt, hex_data[offset:offset+2])[0]
                    offset += 2
                    if offset + str_len > total_length or str_len == 0:
                        break

                    string_data = hex_data[offset:offset+str_len]
                    offset += str_len
                    encoding = encoding_map.get(string_code, 'ascii')

                    if encoding == 'hex_ascii':
                        hex_str = string_data.hex() if msb_first else string_data[::-1].hex()
                        result = f"0x{hex_str.upper()}"
                    else:
                        result = string_data.decode(encoding, errors='ignore')
                        group_part = result

            elif is_raw:
                if offset + 2 > total_length:
                    break
                raw_len = struct.unpack(str_len_fmt, hex_data[offset:offset+2])[0]
                offset += 2
                if offset + raw_len > total_length or raw_len == 0:
                    break

                raw_data = hex_data[offset:offset+raw_len]
                offset += raw_len
                result = raw_data.hex(' ', 1)
                group_part = f"{{arg.{(_arg_idx + 1):02d}, raw}}"

            elif is_sint or is_usint or is_float or is_bool:
                data_size = data_size_map.get(data_type_length, 0)
                if data_size == 0 or offset + data_size > total_length:
                    break

                data_bytes = hex_data[offset:offset+data_size]
                offset += data_size
                fmt_prefix = '>' if msb_first else '<'
                fmt = ""

                if is_sint:
                    fmt = f"{fmt_prefix}{'bhi'[data_size // 2]}" if data_size in (1,2,4) else ""
                elif is_usint:
                    fmt = f"{fmt_prefix}{'BHI'[data_size // 2]}" if data_size in (1,2,4) else ""

                is_hex_format = False
                if fmt:
                    try:
                        result = str(struct.unpack(fmt, data_bytes)[0])
                    except struct.error:
                        result = ""

                if is_usint and string_code == 0x02 and result:
                    hex_str = data_bytes.hex() if msb_first else data_bytes[::-1].hex()
                    result = f"0x{hex_str.upper()}"
                    is_hex_format = True

                if is_bool:
                    datatype = 'bool'
                elif is_sint:
                    datatype = f"sint{ {1:'8',2:'16',4:'32',8:'64'}[data_size] }"
                elif is_usint:
                    datatype = f"uint{ {1:'8',2:'16',4:'32',8:'64'}[data_size] }"
                    if is_hex_format:
                        datatype += "_hexstr"
                elif is_float:
                    datatype = f"float{ {4:'32',8:'64'}[data_size] }"
                else:
                    datatype = ''

                group_part = f"{{arg.{(_arg_idx + 1):02d}, {datatype}}}" if datatype else result
            else:
                result = hex_data[offset:].hex(' ', 1) if offset < total_length else ""
                offset = total_length
                group_part = result

            payload_parts.append(result)
            payload_p_parts.append(group_part)

        payload = " ".join(payload_parts).replace('\x00', '').strip()
        payload_p = " ".join(payload_p_parts).replace('\x00', '').strip()
        return payload, payload_p

    @classmethod
    def _parse_non_verbose_payload(cls, hex_data):
        try:
            if len(hex_data) < cls.NON_VERBOSE_MSG_ID_LEN:
                message_id_str = "[0]"
                data_bytes = b""
            else:
                message_id = struct.unpack('>I', hex_data[:cls.NON_VERBOSE_MSG_ID_LEN])[0]
                message_id_str = f"[{message_id}]"
                data_bytes = hex_data[cls.NON_VERBOSE_MSG_ID_LEN:]
                separator_idx = data_bytes.find(cls.DLT_SEPARATOR)
                if separator_idx != -1:
                    data_bytes = data_bytes[:separator_idx]

            if not data_bytes:
                data_part = ""
            else:
                ascii_str = data_bytes.decode('ascii', errors='ignore')
                hex_str = data_bytes.hex(' ', 1)
                data_part = f"{ascii_str}|{hex_str}"

            payload = f"{message_id_str} {data_part}".strip()
            return payload, payload
        except Exception:
            return "", ""

    @classmethod
    def split_dlt_messages(cls, file_data):
        messages = []
        start = 0
        magic = cls.MAGIC_HEADER
        magic_len = len(magic)
        file_len = len(file_data)

        while start < file_len:
            next_start = file_data.find(magic, start + magic_len)
            if next_start == -1:
                messages.append(file_data[start:])
                break
            messages.append(file_data[start:next_start])
            start = next_start
        return messages

    @classmethod
    def _extract_cal_timestamp(cls, ecuid, payload, timestamp):
        try:
            if ecuid in ["CCU0", "CCU1"]:
                match = cls.CAL_TIMESTAMP_PATTERN.search(payload)
                if match:
                    return round(float(match.group(1)), 4)
            return round(timestamp, 4) if timestamp is not None else 0.0
        except Exception:
            return round(timestamp, 4) if timestamp is not None else 0.0

    @classmethod
    def parse_dlt_message(cls, hex_data, msg_index):
        if len(hex_data) < cls.MIN_MSG_LENGTH:
            return None
        try:
            storage_header = cls.parse_storage_header(hex_data[:cls.STORAGE_HEADER_LEN])
            standard_header = cls.parse_standard_header(hex_data[cls.STORAGE_HEADER_LEN:cls.STORAGE_HEADER_LEN+cls.STANDARD_HEADER_LEN])
            payload_length = standard_header['length'] - cls.STANDARD_HEADER_LEN - cls.EXTENDED_HEADER_LEN
            
            if payload_length < 0:
                return None

            extended_header = cls.parse_extended_header(hex_data[32:42])
            payload_data = hex_data[42:42+payload_length]
            payload, payload_p = cls.parse_payload(payload_data, extended_header['num_arguments'], standard_header['msb_first'], extended_header['mode'])
            cal_timestamp = cls._extract_cal_timestamp(storage_header['ecuid'], payload, standard_header['timestamp'])

            return (
                msg_index, storage_header['time'], standard_header['timestamp'], cal_timestamp,
                standard_header['message_counter'], storage_header['ecuid'],
                extended_header['application_id'], extended_header['context_id'],
                standard_header['session_id'], extended_header['message_type'],
                extended_header['message_type_info'], 'verbose' if extended_header['mode'] else 'non-verbose',
                extended_header['num_arguments'], payload, payload_p
            )
        except Exception:
            return None

RT1_signals = [
    ["", "", "", "NMiDPSNMKeepAwakeFlag", "NMEEMKeepAwakeFlag", "NMEEMKeepAwakeFlag_Aux",
     "NMVCUGlobalKeepAwakeFlag", "NMVCUChargingKeepAwakeFlag(新能源,VCU_hcu)"],
    ["DoIPActiveLineStatus", "NMBMSGlobalKeepAwakeFlag", "NMBMSChargingKeepAwakeFlag",
     "NMBMSThermalRunawayKeepAwakeFlag(新能源,BMS)", "-", "NMBMSThermalRunawayKeepAwakeFlag(动力,VCU_slccu)",
     "NMVCUChargingKeepAwakeFlag(动力,VCU_slccu)", "NMVCU2GlobalKeepAwakeFlag"],
    ["NMVHMKeepAwakeFlag", ".", "CSM_RC_KeepAwake", "NMHVACKeepAwakeFlag(新能源,TMS)", ".", ".",
     "BDC_NMBDCKeepAwakeFlag", "NMTMSChargingKeepAwakeFlag"],
    ["NMTMSGlobalKeepAwakeFlag", "NMHMPNCDownloadKeepAwakeFlag", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    ["NMWCCDownloadRequest_RT1", "CAN7_DiagMSG_ID ！= 0x74F（PNCGlobal）", "CAN7_DiagMSG_ID == 0x74F（RequestOBD）", ".",
     ".", ".", ".", "."],
    ["SPME触发A核网络维持", "SMM触发A核网络维持", "CSM触发A核网络维持", "OTACLT触发A核网络维持", "VCP触发A核网络维持",
     "BDC触发A核网络维持", "VT触发A核网络维持", "."],
    [".", ".", ".", ".", ".", ".", ".", "."]
]

Event_Source_signals = ["PNCGlobal", "PNCCharging", "PNCBSM", "PNCDownload", "PNCEnter", "PNCHazard", "Reserved", "PNCRearlight"]
Acore_signals = ["SPME(PNCGlobal)", "SMM(PNCGlobal)", "CSM(PNCGlobal)", "OTACLT(PNCDownload)", "VCP(PNCGlobal)",
                 "BDC(PNCDownload,PNCGlobal)", "VT(PNCGlobal)", "UA"]

RT2_signals = [
    ["NMATWSKeepAwakeFlag", "NMExteriorLightKeepAwakeFlag", "NMLockingKeepAwakeFlag", "PWL_NMPWLKeepAwakeFlag",
     "SR_NMSunRoofKeepAwakeFlag", "NMPLGKeepAwakeFlag", "NMPowerOperatedDoorKeepAwakeFlag", "MR_NMMirrorKeepAwakeFlag"],
    ["SSW_NMSeatControlKeepAwakeFlag", "NMHVACKeepAwakeFlag", "NMHVACRLSLinAwakeFlag", "NMAQCKeepAwakeFlag",
     "NMTPMSKeepAwakeFlag", "ILC_NMKeepAwakeFlag", "NMESMKeepAwakeFlag", "NMKAEKeepAwakeFlag"],
    ["NMVRMPNCCharingKeepAwakeFlag", "SMM_NMLPDKeepAwakeFlag", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    ["NMWCCDownloadRequest_RT2", ".", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."],
    [".", ".", ".", ".", ".", ".", ".", "."]
]

def rt1_detect_and_modify(txt_file):
    keyword = ["RT1_McoreSignalChange", "RT1_NetWUStVar"]
    df = pd.read_csv(txt_file, delimiter=r'\s+')
    ninth_column = df.iloc[:, 13]
    keyword_indices = ninth_column[ninth_column.str.contains('|'.join(keyword))].index
    for index in keyword_indices:
        ans = "--"
        temp_str = df.iloc[index, 13]
        numbers_string = temp_str.split("):")[1]
        numbers_list = numbers_string.split(",")[:8]
        bin_array = []
        for dec in numbers_list:
            decimal_value = int(dec, 16)
            binary_string = bin(decimal_value)[2:]
            padded_binary_str = binary_string.zfill(8)
            bin_array.append(padded_binary_str)
        car_state = bin_array[0][5] + bin_array[0][6] + bin_array[0][7]
        if car_state == "000":
            ans += "Standby--"
        elif car_state == "001":
            ans += "Comfort--"
        elif car_state == "011":
            ans += "DRV--"
        for i in range(len(bin_array)):
            for j in range(8):
                if bin_array[i][j] == "1":
                    ans += RT1_signals[i][7 - j]
                    ans += "--"
        df.iloc[index, 13] += f"{ans}"
    df.to_csv(txt_file, sep=' ', index=False)

def rt2_detect_and_modify(txt_file):
    try:
        keyword = ["RT2_McoreSignalChange", "RT2_NetWUStVar"]
        df = pd.read_csv(txt_file, delimiter=r'\s+', dtype=str)
        ninth_column = df.iloc[:, 13]
        keyword_indices = ninth_column[ninth_column.str.contains('|'.join(keyword))].index
        for index in keyword_indices:
            ans = "--"
            temp_str = df.iloc[index, 13]
            numbers_string = temp_str.split("):")[1]
            numbers_list = numbers_string.split(",")[:8]
            bin_array = []
            for dec in numbers_list:
                decimal_value = int(dec, 16)
                binary_string = bin(decimal_value)[2:]
                padded_binary_str = binary_string.zfill(8)
                bin_array.append(padded_binary_str)
            for i in range(len(bin_array)):
                for j in range(8):
                    if bin_array[i][j] == "1":
                        ans += RT2_signals[i][7 - j]
                        ans += "--"
            df.iloc[index, 13] += f"{ans}"
        df.to_csv(txt_file, sep=' ', index=False)
    except Exception as e:
        print(f"处理文件 {txt_file} 时出现错误：{str(e)}")

def eventSource_current_detect_and_modify(txt_file):
    keyword = ["EventSource_Current", "EvenSource_Current"]
    df = pd.read_csv(txt_file, delimiter=r'\s+')
    ninth_column = df.iloc[:, 13]
    keyword_indices = ninth_column[ninth_column.str.contains('|'.join(keyword))].index
    for index in keyword_indices:
        ans = "--"
        temp_str = df.iloc[index, 13]
        numbers_string = temp_str.split("):")[1]
        numbers_list = numbers_string.split(",")[:1]
        bin_array = []
        for dec in numbers_list:
            decimal_value = int(dec, 16)
            binary_string = bin(decimal_value)[2:]
            padded_binary_str = binary_string.zfill(8)
            bin_array.append(padded_binary_str)
        for i in range(len(bin_array)):
            for j in range(8):
                if bin_array[i][j] == "1":
                    ans += Event_Source_signals[7 - j]
                    ans += "--"
        df.iloc[index, 13] += f"{ans}"
    df.to_csv(txt_file, sep=' ', index=False)

def acore_signal_detect_and_modify(txt_file):
    keyword = "AcoreSignalChange"
    df = pd.read_csv(txt_file, delimiter=r'\s+')
    ninth_column = df.iloc[:, 13]
    keyword_indices = ninth_column[ninth_column.str.contains(keyword)].index
    for index in keyword_indices:
        ans = "--"
        temp_str = df.iloc[index, 13]
        numbers_string = temp_str.split("):")[1]
        numbers_list = numbers_string.split(",")[:1]
        bin_array = []
        for dec in numbers_list:
            decimal_value = int(dec, 16)
            binary_string = bin(decimal_value)[2:]
            padded_binary_str = binary_string.zfill(8)
            bin_array.append(padded_binary_str)
        for i in range(len(bin_array)):
            for j in range(8):
                if bin_array[i][j] == "1":
                    ans += Acore_signals[7 - j]
                    ans += "--"
        df.iloc[index, 13] += f"{ans}"
    df.to_csv(txt_file, sep=' ', index=False)

def ensure_dlt_extension(filename):
    if '.' not in filename:
        new_filename = filename + '.dlt'
    else:
        base_name, extension = os.path.splitext(filename)
        if extension != '.dlt':
            new_filename = base_name + '.dlt'
        else:
            new_filename = filename
    if new_filename == filename:
        return filename
    try:
        os.rename(filename, new_filename)
        return new_filename
    except OSError as e:
        print(f'Error: {e}')
        return filename

def parse_single_message(args):
    msg, idx = args
    return LogParser.parse_dlt_message(msg, idx)

def convert_dlt_to_txt(input_dlt_file, output_txt_file, max_workers=4):
    try:
        with open(input_dlt_file, 'rb') as f:
            file_data = f.read()
        
        messages = LogParser.split_dlt_messages(file_data)
        parsed_results = []
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            args_list = [(msg, idx) for idx, msg in enumerate(messages)]
            results = list(executor.map(parse_single_message, args_list))
            
        for res in results:
            if res:
                filtered_row = list(res[:3]) + list(res[4:14])
                parsed_results.append(filtered_row)
        
        with open(output_txt_file, 'w', encoding='utf-8') as f:
            for row in parsed_results:
                row_str = " ".join([str(item).replace('\t', ' ').replace('\x00', '').strip() for item in row])
                f.write(row_str + "\n")
        
        print(f"DLT文件 {input_dlt_file} 原生转换为TXT成功！")
        return 0
    except Exception as e:
        print(f"DLT转换失败：{e}")
        return -1

def txt_to_excel(txt_file, excel_file):
    df = pd.read_csv(txt_file, sep=' ', header=None)
    df.to_excel(excel_file, index=False, header=False, engine='openpyxl')
    print("转换完成，Excel 文件保存在:", excel_file)

def decompress_folder(folder):
    temp_folder = os.path.join(folder, 'temp')
    if not os.path.exists(temp_folder):
        os.makedirs(temp_folder)

    for file in os.listdir(folder):
        if file.endswith('.tar.gz'):
            tar_gz_file_path = os.path.join(folder, file)
            decompressed_folder_path = os.path.join(temp_folder, file[:-7])
            try:
                with tarfile.open(tar_gz_file_path, 'r:gz') as tar:
                    tar.extractall(path=decompressed_folder_path)
                os.remove(tar_gz_file_path)
                print(f"已解压并删除: {tar_gz_file_path}")
            except Exception as e:
                print(f"处理文件 {tar_gz_file_path} 失败: {e}")

    for root, dirs, files in os.walk(folder, topdown=False):
        for file in files:
            file_path = os.path.join(root, file)
            if 'WCC' in file or 'POWM' in file:
                target_path = os.path.join(folder, file)
                shutil.move(file_path, target_path)
                print(f"保留文件: {file_path} 移动到: {target_path}")
            else:
                os.remove(file_path)
                print(f"删除文件: {file_path}")
        for dir in dirs:
            dir_path = os.path.join(root, dir)
            os.rmdir(dir_path)
            print(f"删除文件夹: {dir_path}")

    files = os.listdir(folder)
    for file in files:
        if file.endswith('.gz'):
            gzipped_file = os.path.join(folder, file)
            output_file = os.path.join(folder, os.path.splitext(file)[0])
            if os.path.exists(output_file + ".dlt"):
                print(f"已存在解压后的文件: {output_file}")
                continue
            with gzip.open(gzipped_file, 'rb') as f_in:
                with open(output_file, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            print(f"解压缩 {gzipped_file} 到 {output_file}")

class StreamToTextEdit(io.StringIO):
    def __init__(self, text_edit):
        super().__init__()
        self.text_edit = text_edit
    def write(self, text):
        self.text_edit.insertPlainText(text)
        self.text_edit.ensureCursorVisible()

class MyWorkerThread(QThread):
    progress_signal = pyqtSignal(int)
    finished_signal = pyqtSignal()
    log_data_signal = pyqtSignal(list)

    def __init__(self, folder_path, parent=None):
        super().__init__(parent)
        self.folder_path = folder_path

    def run(self):
        start_time = time.time()
        print(f"程序开始于:🕑 {time.ctime(start_time)}")
        workbook = openpyxl.Workbook()
        folder_path = self.folder_path
        current_step = 0
        total_steps = 11

        decompress_folder(folder_path)
        files = os.listdir(folder_path)
        input_dlt_file = [file for file in files if file.endswith('.dlt') or file.endswith('.log') or '.' not in file]
        for i in range(0, len(input_dlt_file)):
            input_dlt_file[i] = ensure_dlt_extension(folder_path + "/" + input_dlt_file[i])
        output_txt_file = [f"{os.path.splitext(file)[0]}.txt" for file in input_dlt_file]
        all_output_txt_file = folder_path + "/AAAAAAAAALLLL.txt"

        now = datetime.now()
        time_string = now.strftime("%Y-%m-%d_%H:%M:%S").replace(" ", "_").replace(":", "_")
        output_excel = folder_path + "/" + time_string + ".xlsx"
        workbook.save(output_excel)

        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        for dlt, txt in zip(input_dlt_file, output_txt_file):
            convert_dlt_to_txt(dlt, txt)

        try:
            with open(all_output_txt_file, 'w', encoding='utf-8') as outfile:
                for file_name in output_txt_file:
                    with open(file_name, 'r', encoding='utf-8') as infile:
                        outfile.write(infile.read())
            print(f"合并完成，结果保存在 {all_output_txt_file}")
        except Exception as e:
            print("合并过程中出现错误:", e)

        for file_path in output_txt_file:
            if os.path.exists(file_path):
                os.remove(file_path)

        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        with open(all_output_txt_file, 'r+') as file:
            lines = file.readlines()
            for index, line in enumerate(lines):
                columns = line.strip().split()
                if len(columns) < 14:
                    continue
                if "DPT_VER" in columns[13]:
                    columns[13] = "DPT_VER_illegal"
                elif "DSPEC_VER" in columns[13]:
                    columns[13] = "DSPEC_VER_illegal"
                lines[index] = ' '.join(columns) + '\n'
            file.seek(0)
            file.writelines(lines)

        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        with open(all_output_txt_file, 'r+') as f:
            lines = f.readlines()
            for i, line in enumerate(lines):
                space_count = 0
                index = 0
                for j, char in enumerate(line):
                    if char == ' ':
                        space_count += 1
                        if space_count == 13:
                            index = j + 1
                            break
                if index > 0:
                    modified_line = line[:index] + line[index:].replace(' ', '_')
                    lines[i] = modified_line
            f.seek(0)
            f.writelines(lines)
            f.truncate()

        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        with open(all_output_txt_file, 'r+') as file:
            lines = file.readlines()
            for i, line in enumerate(lines):
                parts = line.strip().split()
                if len(parts) >= 3:
                    combined = parts[1] + '__' + parts[2]
                    parts[1] = combined
                    lines[i] = ' '.join(parts) + '\n'
            file.seek(0)
            file.writelines(lines)
            file.truncate()

        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        rt1_detect_and_modify(all_output_txt_file)
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))
        rt2_detect_and_modify(all_output_txt_file)
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))
        eventSource_current_detect_and_modify(all_output_txt_file)
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))
        acore_signal_detect_and_modify(all_output_txt_file)
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        txt_to_excel(all_output_txt_file, output_excel)
        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))

        log_data = []
        with open(all_output_txt_file, 'r', encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split()
                if len(parts) >= 14:
                    log_data.append(parts[:14])
        
        self.log_data_signal.emit(log_data)

        os.remove(all_output_txt_file)

        header = ['Index', 'Date', 'Time', 'TimeStamp', 'Count', 'EcuID','Apid', 'Ctid', 'sessionId', 'Type', 'Subtype', 'Mode','Args', 'PayLoad']
        workbook = openpyxl.load_workbook(output_excel)
        worksheet = workbook.active
        worksheet.insert_rows(1)
        for col_idx, col_title in enumerate(header, start=1):
            worksheet.cell(row=1, column=col_idx, value=col_title)
        bold_font = Font(bold=True)
        for col_idx in range(1, len(header) + 1):
            worksheet.cell(row=1, column=col_idx).font = bold_font
        worksheet.freeze_panes = 'A2'
        columns_to_hide = ['C', 'H', 'I', 'J', 'K', 'L', 'M']
        for col in columns_to_hide:
            worksheet.column_dimensions[col].hidden = True
        worksheet.column_dimensions['B'].width = 32
        worksheet.column_dimensions['D'].width = 13

        darkorange = PatternFill(start_color="E26B0A", end_color="00008B", fill_type="solid")
        red = PatternFill(start_color="FFD699", end_color="FF0000", fill_type="solid")
        dark_blue = PatternFill(start_color="6495ED", end_color="00008B", fill_type="solid")
        qianqing = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")

        for row in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=14)
            if "RT1_O" in str(cell.value):
                cell.fill = red
            elif "RT2_O" in str(cell.value):
                cell.fill = qianqing
            elif "RT1_M" in str(cell.value):
                cell.fill = darkorange
            elif "RT2_M" in str(cell.value):
                cell.fill = dark_blue

        workbook.save(output_excel)

        current_step += 1
        self.progress_signal.emit(int((current_step / total_steps) * 100))
        end_time = time.time()
        print(f"程序结束于:⏰ {time.ctime(end_time)}")
        print(f"程序总运行时间:⏳ {end_time - start_time:.2f}秒")
        self.finished_signal.emit()

class LogViewerWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.all_log_data = []
        self.first_timestamp = None
        self.initUI()
        
    def initUI(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        
        filter_layout = QHBoxLayout()
        
        filter_layout.addWidget(QLabel("开始时间:"))
        self.start_time_edit = QDateTimeEdit(self)
        self.start_time_edit.setCalendarPopup(True)
        self.start_time_edit.setDisplayFormat("yyyy/MM/dd HH:mm:ss")
        self.start_time_edit.setDateTime(QDateTime.currentDateTime().addDays(-1))
        filter_layout.addWidget(self.start_time_edit)
        
        filter_layout.addWidget(QLabel("结束时间:"))
        self.end_time_edit = QDateTimeEdit(self)
        self.end_time_edit.setCalendarPopup(True)
        self.end_time_edit.setDisplayFormat("yyyy/MM/dd HH:mm:ss")
        self.end_time_edit.setDateTime(QDateTime.currentDateTime())
        filter_layout.addWidget(self.end_time_edit)
        
        self.sort_combo = QComboBox(self)
        self.sort_combo.addItems(["按时间升序", "按时间降序"])
        filter_layout.addWidget(QLabel("排序:"))
        filter_layout.addWidget(self.sort_combo)
        
        self.filter_btn = QPushButton("筛选", self)
        self.filter_btn.clicked.connect(self.apply_filter)
        filter_layout.addWidget(self.filter_btn)
        
        self.reset_btn = QPushButton("重置", self)
        self.reset_btn.clicked.connect(self.reset_filter)
        filter_layout.addWidget(self.reset_btn)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        self.log_table = QTableWidget(self)
        self.log_table.setColumnCount(5)
        self.log_table.setHorizontalHeaderLabels(["时间", "相对时间(s)", "EcuID", "Application ID", "Payload"])
        self.log_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.log_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.log_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.log_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.log_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self.log_table.setAlternatingRowColors(True)
        self.log_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.log_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        self.log_table.setStyleSheet("""
            QTableWidget {
                background-color: #FAFAFA;
                alternate-background-color: #F0F8FF;
                gridline-color: #D0D0D0;
                border: 1px solid #A0A0A0;
            }
            QTableWidget::item {
                padding: 2px;
            }
            QHeaderView::section {
                background-color: #4A90D9;
                color: white;
                padding: 4px;
                border: 1px solid #357ABD;
                font-weight: bold;
            }
        """)
        
        layout.addWidget(self.log_table)
        
        self.status_label = QLabel("共 0 条记录", self)
        layout.addWidget(self.status_label)
        
    def parse_time_string(self, time_str):
        try:
            if '__' in time_str:
                datetime_part = time_str.split('__')[0]
            else:
                datetime_part = time_str
            return datetime.strptime(datetime_part, '%Y/%m/%d %H:%M:%S.%f')
        except:
            return None
            
    def set_log_data(self, log_data):
        self.all_log_data = log_data
        if log_data:
            first_time = self.parse_time_string(log_data[0][1] if len(log_data[0]) > 1 else "")
            if first_time:
                self.first_timestamp = first_time.timestamp()
            else:
                self.first_timestamp = None
        else:
            self.first_timestamp = None
        self.display_logs(log_data)
        
    def display_logs(self, logs):
        self.log_table.setRowCount(0)
        self.log_table.setRowCount(len(logs))
        
        for row, log in enumerate(logs):
            if len(log) >= 14:
                time_str = log[1] if len(log) > 1 else ""
                time_item = QTableWidgetItem(time_str)
                time_item.setData(Qt.UserRole, self.parse_time_string(time_str))
                self.log_table.setItem(row, 0, time_item)
                
                relative_time = ""
                if self.first_timestamp:
                    try:
                        current_time = self.parse_time_string(time_str)
                        if current_time:
                            relative_time = f"{current_time.timestamp() - self.first_timestamp:.4f}"
                    except:
                        pass
                self.log_table.setItem(row, 1, QTableWidgetItem(relative_time))
                
                ecuid = log[5] if len(log) > 5 else ""
                self.log_table.setItem(row, 2, QTableWidgetItem(ecuid))
                
                app_id = log[6] if len(log) > 6 else ""
                self.log_table.setItem(row, 3, QTableWidgetItem(app_id))
                
                payload = log[13] if len(log) > 13 else ""
                self.log_table.setItem(row, 4, QTableWidgetItem(payload))
                
        self.status_label.setText(f"共 {len(logs)} 条记录")
        
    def apply_filter(self):
        start_dt = self.start_time_edit.dateTime().toPyDateTime()
        end_dt = self.end_time_edit.dateTime().toPyDateTime()
        
        filtered_logs = []
        for log in self.all_log_data:
            if len(log) >= 14:
                time_str = log[1] if len(log) > 1 else ""
                log_time = self.parse_time_string(time_str)
                if log_time and start_dt <= log_time <= end_dt:
                    filtered_logs.append(log)
        
        if self.sort_combo.currentIndex() == 1:
            filtered_logs.sort(key=lambda x: self.parse_time_string(x[1] if len(x) > 1 else "") or datetime.min, reverse=True)
        else:
            filtered_logs.sort(key=lambda x: self.parse_time_string(x[1] if len(x) > 1 else "") or datetime.min)
            
        self.display_logs(filtered_logs)
        
    def reset_filter(self):
        self.start_time_edit.setDateTime(QDateTime.currentDateTime().addDays(-1))
        self.end_time_edit.setDateTime(QDateTime.currentDateTime())
        self.sort_combo.setCurrentIndex(0)
        self.display_logs(self.all_log_data)

class DltConverter(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("DLT文件转换与日志查看器")
        self.setGeometry(300, 100, 1100, 800)
        
        self.setStyleSheet("""
            QMainWindow {
                background-color: #F5F5F5;
            }
            QLabel {
                font-family: 'Microsoft YaHei';
                font-size: 12px;
            }
            QPushButton {
                background-color: #4A90D9;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-family: 'Microsoft YaHei';
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #357ABD;
            }
            QPushButton:pressed {
                background-color: #2A5F8F;
            }
            QPushButton:disabled {
                background-color: #A0A0A0;
            }
            QLineEdit {
                padding: 6px;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                background-color: white;
            }
            QProgressBar {
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                text-align: center;
                background-color: #E0E0E0;
            }
            QProgressBar::chunk {
                background-color: #4A90D9;
                border-radius: 3px;
            }
            QTabWidget::pane {
                border: 1px solid #CCCCCC;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #E0E0E0;
                padding: 8px 16px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: #4A90D9;
                color: white;
            }
        """)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        file_layout = QHBoxLayout()
        self.label_dlt_file = QLabel("DLT文件夹:")
        file_layout.addWidget(self.label_dlt_file)
        self.entry_dlt_file = QLineEdit(self)
        self.entry_dlt_file.setMinimumWidth(400)
        file_layout.addWidget(self.entry_dlt_file)
        self.button_select_dlt_file = QPushButton("选择文件夹", self)
        self.button_select_dlt_file.clicked.connect(self.select_dlt_file)
        file_layout.addWidget(self.button_select_dlt_file)
        file_layout.addStretch()
        main_layout.addLayout(file_layout)

        self.label_instructions = QLabel(
            "使用说明：\n"
            "1. 将要转换的日志文件放入同一个文件夹\n"
            "2. 日志文件可以是.gz压缩包，也可以是.dlt，也可以是.log\n"
            "3. 转换完成后统统会变为.dlt文件\n"
            "4. 完成后所有日志会被合并进同一个xlsx文件\n"
            "5. 可在「日志查看」标签页查看、筛选和排序日志"
        )
        self.label_instructions.setStyleSheet("color: #333333; font-family: 'Microsoft YaHei'; background-color: #E8F4FD; padding: 10px; border-radius: 5px;")
        main_layout.addWidget(self.label_instructions)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        self.button_convert = QPushButton("开始转换", self)
        self.button_convert.setMinimumWidth(150)
        self.button_convert.clicked.connect(self.convert)
        btn_layout.addWidget(self.button_convert)
        btn_layout.addStretch()
        main_layout.addLayout(btn_layout)

        self.progress_bar = QProgressBar(self)
        self.progress_bar.setValue(0)
        main_layout.addWidget(self.progress_bar)

        self.tab_widget = QTabWidget(self)
        
        self.log_text_edit = QTextEdit(self)
        self.log_text_edit.setReadOnly(True)
        self.log_text_edit.setStyleSheet("""
            QTextEdit {
                background-color: #1E1E1E;
                color: #00FF00;
                border: 2px solid #333333;
                border-radius: 5px;
                padding: 10px;
                font-family: 'Consolas', 'Microsoft YaHei';
                font-size: 12px;
            }
        """)
        self.tab_widget.addTab(self.log_text_edit, "运行日志")
        
        self.log_viewer = LogViewerWidget(self)
        self.tab_widget.addTab(self.log_viewer, "日志查看")
        
        main_layout.addWidget(self.tab_widget)
        
        sys.stdout = StreamToTextEdit(self.log_text_edit)

    def select_dlt_file(self):
        file_path = QFileDialog.getExistingDirectory(self, "选择DLT文件夹")
        if file_path:
            self.entry_dlt_file.setText(file_path)

    def convert(self):
        folder_path = self.entry_dlt_file.text()
        if not os.path.isdir(folder_path):
            QMessageBox.critical(self, "错误", "请选择有效的DLT文件夹")
            return
        self.button_convert.setEnabled(False)
        self.worker_thread = MyWorkerThread(folder_path)
        self.worker_thread.progress_signal.connect(self.update_progress)
        self.worker_thread.finished_signal.connect(self.on_conversion_finished)
        self.worker_thread.log_data_signal.connect(self.on_log_data_ready)
        self.worker_thread.start()

    def update_progress(self, progress):
        self.progress_bar.setValue(progress)

    def on_log_data_ready(self, log_data):
        self.log_viewer.set_log_data(log_data)
        self.tab_widget.setCurrentIndex(1)

    def on_conversion_finished(self):
        QMessageBox.information(self, "提示", "转换完成，xlsx文件已保存在dlt文件相同路径的文件夹下。")
        self.button_convert.setEnabled(True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = DltConverter()
    window.show()
    sys.exit(app.exec_())
