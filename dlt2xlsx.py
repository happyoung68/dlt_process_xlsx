import base64
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QMessageBox, QFileDialog, QProgressBar, QTextEdit
from PyQt5.QtCore import QSettings, QThread, pyqtSignal
from PyQt5.QtWidgets import QMainWindow, QLabel, QLineEdit, QApplication
from PyQt5.QtGui import QPixmap, QIcon
import pandas as pd
from datetime import datetime
import os
import subprocess
import openpyxl
from openpyxl.styles import Font, Color, Alignment
import gzip
import shutil
import time
import io

RT1_signals = [
]
Event_Source_signals = ["a", "b", ".", ".", ".", ".", ".", "."]
Acore_signals = [".(PNCGlobal)", ".", "."]
RT2_signals = [
    ["a", "b", ".",".", ".", ".", ".", "."],
    ["." , ".", ".", "." , ".", ".", ".", "."]
]

def rt1_detect_and_modify(txt_file):
    # 读取TXT文件，使用空格作为分隔符
    keyword = ["RT1_McoreSignalChange", "RT1_NetWUStVar"]
    df = pd.read_csv(txt_file, delimiter='\s+')

    # 检测第九列的数据是否包含关键字
    ninth_column = df.iloc[:, 13]  # 第九列的索引是8，因为索引从0开始

    # print(ninth_column.dtype)
    keyword_indices = ninth_column[ninth_column.str.contains('|'.join(keyword))].index

    # 对检测到关键字的行进行操作并添加内容
    for index in keyword_indices:
        # 在第九列后面添加新内容
        ans = "--"
        temp_str = df.iloc[index, 13]

        # 提取后面八个数字
        numbers_string = temp_str.split("):")[1]  # 分割字符串并取第二部分
        numbers_list = numbers_string.split(",")[:8]  # 将字符串分割成数字列表并取前八个数字
        # print(numbers_list)
        bin_array = []
        for dec in numbers_list:
            decimal_value = int(dec, 16)  # 将十六进制字符串转换为十进制数值
            binary_string = bin(decimal_value)[2:]  # 将十进制数值转换为二进制字符串，并去除开头的 '0b'
            padded_binary_str = binary_string.zfill(8)  # 在二进制字符串前面补充0，使其长度为8位
            bin_array.append(padded_binary_str)


        # for i in range(len(bin_array)):
        #     for j in range(8):
        #         if bin_array[i][j] == "1":
        #             ans += RT1_signals[i][7 - j]
        #             ans += "--"


        car_state = bin_array[0][5] + bin_array[0][6] + bin_array[0][7]
        if car_state=="000":
            ans += "Standby--"
        elif car_state=="001":
            ans += "Comfort--"
        elif car_state=="011":
            ans += "DRV--"
        for i in range(len(bin_array)):
            for j in range(8):
                if bin_array[i][j] == "1":
                    ans += RT1_signals[i][7 - j]
                    ans += "--"

        df.iloc[index, 13] += f"{ans}"

    # 将修改后的数据写回TXT文件
    df.to_csv(txt_file, sep=' ', index=False)

def rt2_detect_and_modify(txt_file):
    try:
        # 读取TXT文件，使用空格作为分隔符
        keyword = ["RT2_McoreSignalChange", "RT2_NetWUStVar"]
        df = pd.read_csv(txt_file, delimiter='\s+', dtype=str)

        # 检测第九列的数据是否包含关键字
        ninth_column = df.iloc[:, 13]  # 第九列的索引是8，因为索引从0开始
        # print(ninth_column.dtype)
        keyword_indices = ninth_column[ninth_column.str.contains('|'.join(keyword))].index

        # 对检测到关键字的行进行操作并添加内容
        for index in keyword_indices:
            # 在第九列后面添加新内容
            ans = "--"
            temp_str = df.iloc[index, 13]
            # 提取后面八个数字
            numbers_string = temp_str.split("):")[1]  # 分割字符串并取第二部分
            numbers_list = numbers_string.split(",")[:8]  # 将字符串分割成数字列表并取前八个数字
            # print(numbers_list)
            bin_array = []
            for dec in numbers_list:
                decimal_value = int(dec, 16)  # 将十六进制字符串转换为十进制数值
                binary_string = bin(decimal_value)[2:]  # 将十进制数值转换为二进制字符串，并去除开头的 '0b'
                padded_binary_str = binary_string.zfill(8)  # 在二进制字符串前面补充0，使其长度为8位
                bin_array.append(padded_binary_str)
            for i in range(len(bin_array)):
                for j in range(8):
                    if bin_array[i][j] == "1":
                        ans += RT2_signals[i][7 - j]
                        ans += "--"
            df.iloc[index, 13] += f"{ans}"

        # 将修改后的数据写回TXT文件
        df.to_csv(txt_file, sep=' ', index=False)
    except FileNotFoundError:
        print(f"错误：文件 {txt_file} 未找到。")
    except Exception as e:
        print(f"处理文件 {txt_file} 时出现错误：{str(e)}")

def eventSource_current_detect_and_modify(txt_file):
    # 读取TXT文件，使用空格作为分隔符
    keyword = ["EventSource_Current", "EvenSource_Current"]
    df = pd.read_csv(txt_file, delimiter='\s+')

    # 检测第九列的数据是否包含关键字
    ninth_column = df.iloc[:, 13]  # 第九列的索引是8，因为索引从0开始
    keyword_indices = ninth_column[ninth_column.str.contains('|'.join(keyword))].index

    # 对检测到关键字的行进行操作并添加内容
    for index in keyword_indices:
        # 在第九列后面添加新内容
        ans = "--"
        temp_str = df.iloc[index, 13]
        # 提取后面八个数字
        numbers_string = temp_str.split("):")[1]  # 分割字符串并取第二部分
        numbers_list = numbers_string.split(",")[:1]  # 将字符串分割成数字列表并取前八个数字

        bin_array = []
        for dec in numbers_list:
            decimal_value = int(dec, 16)  # 将十六进制字符串转换为十进制数值
            binary_string = bin(decimal_value)[2:]  # 将十进制数值转换为二进制字符串，并去除开头的 '0b'
            padded_binary_str = binary_string.zfill(8)  # 在二进制字符串前面补充0，使其长度为8位
            bin_array.append(padded_binary_str)
        # print(bin_array)
        for i in range(len(bin_array)):
            for j in range(8):
                if bin_array[i][j] == "1":
                    ans += Event_Source_signals[7 - j]
                    ans += "--"
        df.iloc[index, 13] += f"{ans}"

    # 将修改后的数据写回TXT文件
    df.to_csv(txt_file, sep=' ', index=False)

def acore_signal_detect_and_modify(txt_file):
    # 读取TXT文件，使用空格作为分隔符
    keyword = "AcoreSignalChange"
    df = pd.read_csv(txt_file, delimiter='\s+')

    # 检测第九列的数据是否包含关键字
    ninth_column = df.iloc[:, 13]  # 第九列的索引是8，因为索引从0开始
    keyword_indices = ninth_column[ninth_column.str.contains(keyword)].index

    # 对检测到关键字的行进行操作并添加内容
    for index in keyword_indices:
        # 在第九列后面添加新内容
        ans = "--"
        temp_str = df.iloc[index, 13]
        # 提取后面八个数字
        numbers_string = temp_str.split("):")[1]  # 分割字符串并取第二部分
        numbers_list = numbers_string.split(",")[:1]  # 将字符串分割成数字列表并取前八个数字
        bin_array = []
        for dec in numbers_list:
            decimal_value = int(dec, 16)  # 将十六进制字符串转换为十进制数值
            binary_string = bin(decimal_value)[2:]  # 将十进制数值转换为二进制字符串，并去除开头的 '0b'
            padded_binary_str = binary_string.zfill(8)  # 在二进制字符串前面补充0，使其长度为8位
            bin_array.append(padded_binary_str)
        for i in range(len(bin_array)):
            for j in range(8):
                if bin_array[i][j] == "1":
                    ans += Acore_signals[7 - j]
                    ans += "--"
        df.iloc[index, 13] += f"{ans}"

    # 将修改后的数据写回TXT文件
    df.to_csv(txt_file, sep=' ', index=False)

def ensure_dlt_extension(filename):
    # 检查文件名是否有后缀
    if '.' not in filename:
        new_filename = filename + '.dlt'
    else:
        # 获取文件名和后缀名
        base_name, extension = os.path.splitext(filename)

        # 如果后缀名不是 '.dlt'，则修改为 '.dlt'
        if extension != '.dlt':
            new_filename = base_name + '.dlt'
        else:
            new_filename = filename  # 后缀名已经是 '.dlt'，不需要修改

    # 如果新文件名和旧文件名相同，直接返回
    if new_filename == filename:
        return filename

    # 保存文件（这里假设使用重命名操作来模拟保存）
    try:

        os.rename(filename, new_filename)
        print(f'文件 "{filename}" 重命名为 "{new_filename}"')
        return new_filename
    except OSError as e:
        print(f'FFFError: {e}')
        return filename

def convert_dlt_to_txt(input_dlt_file, output_txt_file, dlt_viewer_path):
    try:
        # 构造调用DLT Viewer的命令
        command = [dlt_viewer_path, '-c', input_dlt_file, output_txt_file]
        # 执行命令
        result = subprocess.run(command, capture_output=True, text=True, check=True)
        # 检查命令执行结果
        if result.returncode == 0:
            print(f"DLT文件 {input_dlt_file} 转换为TXT文件成功！")
        else:
            print("DLT文件转换失败：", result.stderr)
        return result.returncode
    except Exception as e:
        print("DLT文件转换失败：", e)

def txt_to_excel(txt_file, excel_file):
    # 读取文本文件，假设数据以空格分隔
    column_names = ['Index', 'Date', 'Time', 'TimeStamp', 'Count', 'EcuID', 'Apid', 'Ctid', 'sessionId', 'Type', 'Subtype','Mode','Args', 'PayLoad']  # 根据实际情况修改列名
    df = pd.read_csv(txt_file, sep=' ', header=None)
    # print(df)

    # 将 DataFrame 写入 Excel 文件
    df.to_excel(excel_file, index=False, header=False, engine='openpyxl')
    print("转换完成，Excel 文件保存在:", excel_file)

def decompress_folder(folder):
    # 获取文件夹中所有文件的列表
    files = os.listdir(folder)

    # 遍历每个文件
    for file in files:
        if file.endswith('.gz'):
            gzipped_file = os.path.join(folder, file)
            # 输出文件的路径，去掉.gz后缀
            output_file = os.path.join(folder, os.path.splitext(file)[0])

            # 检查输出文件是否已经存在
            if os.path.exists(output_file+".dlt"):
                print(f"已存在解压后的文件: {output_file}")
                continue  # 如果文件已存在，则跳过解压缩

            # 解压缩.gz文件到输出文件
            with gzip.open(gzipped_file, 'rb') as f_in:
                with open(output_file, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)

            print(f"解压缩 {gzipped_file} 到 {output_file}")


# 自定义一个输出流类，用来捕获print的输出
class StreamToTextEdit(io.StringIO):
    def __init__(self, text_edit):
        super().__init__()
        self.text_edit = text_edit

    def write(self, text):
        # 将输出写到QTextEdit控件中
        self.text_edit.insertPlainText(text)
        self.text_edit.ensureCursorVisible()  # 确保光标可见，滚动到底部
class MyWorkerThread(QThread):

    # 定义一个信号，用于传递进度信息
    progress_signal = pyqtSignal(int)
    finished_signal = pyqtSignal()  # 定义一个信号，用于通知任务完成
    dltviewer_error_signal = pyqtSignal()  # 定义一个信号，用于发送错误

    def __init__(self, folder_path, dlt_viewer_path, parent=None):
        super().__init__(parent)
        self.folder_path = folder_path  # 文件夹路径（作为任务的一部分）
        self.dlt_viewer_path = dlt_viewer_path

    def run(self):
        start_time = time.time()
        print(f"程序开始于: {time.ctime(start_time)}")
        # 假设这里是处理文件的过程
        workbook = openpyxl.Workbook()
        folder_path = self.folder_path  # 将self.folder_path赋给局部变量
        current_step = 0
        total_steps = 11  # 假设总共有 11 步

        decompress_folder(folder_path)
        files = os.listdir(folder_path)
        input_dlt_file = [file for file in files if file.endswith('.dlt') or file.endswith('.log') or '.' not in file]
        for i in range(0, len(input_dlt_file)):
            input_dlt_file[i] = ensure_dlt_extension(folder_path + "/" + input_dlt_file[i])
        output_txt_file = []
        for i in range(0, len(input_dlt_file)):
            output_txt_file.append(input_dlt_file[i].split('.')[0] + ".txt")
        for i in range(0, len(output_txt_file)):
            with open(output_txt_file[i], 'w') as file:
                pass
        all_output_txt_file = folder_path + "/AAAAAAAAALLLL.txt"
        with open(all_output_txt_file, 'w') as file:
            pass

        # 获取当前时间
        now = datetime.now()
        # 将当前时间转换为字符串，空格和冒号都替换为下划线
        time_string = now.strftime("%Y-%m-%d_%H:%M:%S").replace(" ", "_").replace(":", "_")
        output_excel = folder_path + "/" + time_string + ".xlsx"
        # output_excel = folder_path + "/AAAALLLLL.xlsx"
        # 保存 Excel 文件
        workbook.save(output_excel)


        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号

        # dlt_viewer_path = self.entry_dlt_viewer.text() + "\dlt_viewer"
        flag = 1
        for i in range(0, len(input_dlt_file)):
            flag = convert_dlt_to_txt(input_dlt_file[i], output_txt_file[i], self.dlt_viewer_path)
        if flag != 0:
            self.dltviewer_error_signal.emit()  #发送错误
            return


        # 合并TXT文件
        try:
            with open(all_output_txt_file, 'w', encoding='utf-8') as outfile:
                for file_name in output_txt_file:
                    with open(file_name, 'r', encoding='utf-8') as infile:
                        outfile.write(infile.read())
            print(f"合并完成，结果保存在 {all_output_txt_file}")
        except Exception as e:
            print("合并过程中出现错误:", e)

        # 删除单个TXT文件
        try:
            for file_path in output_txt_file:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    print(f"临时TXT文件 {file_path} 已经删除.")
                else:
                    print(f"临时TXT文件 {file_path} 不存在.")
        except Exception as e:
            print("删除临时TXT文件失败:", e)


        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号

        # # 合并TXT文件13列后的空格
        # with open(all_output_txt_file, 'r+') as f:
        #     lines = f.readlines()
        #
        #     # 处理每一行数据
        #     for i, line in enumerate(lines):
        #         # 找到第13个空格后的位置
        #         space_count = 0
        #         index = 0
        #         for j, char in enumerate(line):
        #             if char == ' ':
        #                 space_count += 1
        #                 if space_count == 13:
        #                     index = j + 1  # 第13个空格后的位置（即第14个空格的位置）
        #                     break  # 找到位置后跳出循环
        #
        #         # 替换第13个空格后的所有空格为下划线
        #         if index > 0:
        #             modified_line = line[:index] + line[index:].replace(' ', '_')
        #             lines[i] = modified_line  # 更新原始数据
        #
        #     # 将处理后的数据写回文件
        #     f.seek(0)  # 将文件指针移动到文件开头
        #     f.writelines(lines)  # 写入修改后的内容
        #
        #     # 截断文件到当前位置，删除多余内容（如果有）
        #     f.truncate()
        #
        #
        # # 发送进度更新信号，计算进度百分比
        # current_step += 1
        # progress = int((current_step / total_steps) * 100)
        # self.progress_signal.emit(progress)  # 发出进度更新信号

        # 删除TXT文件非法项
        with open(all_output_txt_file, 'r+') as file:
            # 读取文件的所有行
            lines = file.readlines()

            # 遍历每一行进行处理
            for index, line in enumerate(lines):
                # 去除行末的换行符并以空格分割为列表
                columns = line.strip().split()

                # 如果列数小于14，直接跳过
                if len(columns) < 14:
                    continue

                # 检查第14列中的内容并进行相应修改
                if "DPT_VER" in columns[13]:
                    columns[13] = "DPT_VER_illegal"
                elif "DSPEC_VER" in columns[13]:
                    columns[13] = "DSPEC_VER_illegal"

                # 更新修改后的行内容
                lines[index] = ' '.join(columns) + '\n'

            # 将修改后的内容回写到原文件
            file.seek(0)
            file.writelines(lines)

        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号

        #保存一个临时文件
        allthething_txt_file = folder_path + "/allthething.txt"
        # 打开源文件并读取内容
        with open(all_output_txt_file, 'r') as src:
            content = src.read()
        # 将内容写入目标文件
        with open(allthething_txt_file, 'w') as dest:
            dest.write(content)

        # 合并TXT文件13列后的空格
        with open(all_output_txt_file, 'r+') as f:
            lines = f.readlines()

            # 处理每一行数据
            for i, line in enumerate(lines):
                # 找到第13个空格后的位置
                space_count = 0
                index = 0
                for j, char in enumerate(line):
                    if char == ' ':
                        space_count += 1
                        if space_count == 13:
                            index = j + 1  # 第13个空格后的位置（即第14个空格的位置）
                            break  # 找到位置后跳出循环

                # 替换第13个空格后的所有空格为下划线
                if index > 0:
                    modified_line = line[:index] + line[index:].replace(' ', '_')
                    lines[i] = modified_line  # 更新原始数据

            # 将处理后的数据写回文件
            f.seek(0)  # 将文件指针移动到文件开头
            f.writelines(lines)  # 写入修改后的内容

            # 截断文件到当前位置，删除多余内容（如果有）
            f.truncate()

        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号

        # 合并TXT文件时间日期列
        with open(all_output_txt_file, 'r+') as file:
            lines = file.readlines()  # 读取所有行数据

            # 修改每行数据
            for i, line in enumerate(lines):
                parts = line.strip().split()
                if len(parts) >= 3:
                    # 合并第二列和第三列并覆盖第二列
                    combined = parts[1] + '__' + parts[2]
                    parts[1] = combined
                    lines[i] = ' '.join(parts) + '\n'

            # 将文件指针移动到文件开头
            file.seek(0)
            # 写入修改后的所有行数据
            file.writelines(lines)
            # 截断文件，删除多余的内容（如果有）
            file.truncate()


        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号
        rt1_detect_and_modify(all_output_txt_file)
        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号
        rt2_detect_and_modify(all_output_txt_file)
        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号
        eventSource_current_detect_and_modify(all_output_txt_file)
        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号
        acore_signal_detect_and_modify(all_output_txt_file)
        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号

        txt_to_excel(all_output_txt_file, output_excel)

        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号

        # 删除ALL_TXT文件
        try:
            # 删除文件
            os.remove(all_output_txt_file)
            print(f"总文件 {all_output_txt_file} 已删除.")
        except FileNotFoundError:
            print(f"总文件 {all_output_txt_file} 不存在.")
        except Exception as e:
            print("删除总文件失败:", e)

        # # 读取Excel文件
        # df = pd.read_excel(output_excel)
        # df.iloc[:, 1] = df.iloc[:, 1] + " " + df.iloc[:, 2]
        # df.to_excel(output_excel, index=False)
        # update_progress()

        # workbook = openpyxl.load_workbook(output_excel)
        # worksheet = workbook.active
        # worksheet.cell(row=1, column=1).value = 'Index'
        # worksheet.cell(row=1, column=2).value = 'Date'
        # worksheet.cell(row=1, column=3).value = 'Time'
        # worksheet.cell(row=1, column=4).value = 'TimeStamp'
        # worksheet.cell(row=1, column=5).value = 'Count'
        # worksheet.cell(row=1, column=6).value = 'EcuID'
        # worksheet.cell(row=1, column=7).value = 'Apid'
        # worksheet.cell(row=1, column=8).value = 'Ctid'
        # worksheet.cell(row=1, column=9).value = 'sessionId'
        # worksheet.cell(row=1, column=10).value = 'Type'
        # worksheet.cell(row=1, column=11).value = 'Subtype'
        # worksheet.cell(row=1, column=12).value = 'Mode'
        # worksheet.cell(row=1, column=13).value = 'Args'
        # worksheet.cell(row=1, column=14).value = 'PayLoad'
        #
        # cell_range = worksheet['A1:N1']
        # # 设置加粗字体样式
        # bold_font = Font(bold=True)
        # # 对每个单元格应用加粗字体样式
        # for row in cell_range:
        #     for cell in row:
        #         cell.font = bold_font
        # # 冻结单元格
        # worksheet.freeze_panes = 'A2'
        # # 保存修改后的Excel文件
        # workbook.save(output_excel)

        #数据表头
        # header = [
        #     'Index', 'Date', 'Time', 'TimeStamp', 'Count', 'EcuID',
        #     'Apid', 'Ctid', 'sessionId', 'Type', 'Subtype', 'Mode',
        #     'Args', 'PayLoad'
        # ]
        # # 打开现有的 Excel 文件
        # workbook = openpyxl.load_workbook(output_excel)
        # print(111)
        # # 选择要操作的工作表
        # worksheet = workbook.active
        # # 在第一行上方插入一行
        # worksheet.insert_rows(1)
        #
        # # 设置表头
        # for col_idx, col_title in enumerate(header, start=1):
        #     worksheet.cell(row=1, column=col_idx).value = col_title
        # # 设置加粗字体样式
        # bold_font = Font(bold=True)
        # # 对整个表头范围应用加粗字体样式
        # for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(header)):
        #     for cell in row:
        #         cell.font = bold_font
        # # 冻结首行作为标题行
        # worksheet.freeze_panes = 'A2'
        # # 保存修改后的 Excel 文件
        # workbook.save(output_excel)

        header = [
            'Index', 'Date', 'Time', 'TimeStamp', 'Count', 'EcuID',
            'Apid', 'Ctid', 'sessionId', 'Type', 'Subtype', 'Mode',
            'Args', 'PayLoad'
        ]

        # 打开现有的 Excel 文件
        workbook = openpyxl.load_workbook(output_excel)
        print("111处理表格中")
        # 选择要操作的工作表
        worksheet = workbook.active

        # 在第一行上方插入一行
        worksheet.insert_rows(1)

        # 批量设置表头内容
        for col_idx, col_title in enumerate(header, start=1):
            worksheet.cell(row=1, column=col_idx, value=col_title)

        # 设置加粗字体样式（批量设置）
        bold_font = Font(bold=True)
        for col_idx in range(1, len(header) + 1):
            worksheet.cell(row=1, column=col_idx).font = bold_font

        # 冻结首行作为标题行
        worksheet.freeze_panes = 'A2'

        # 保存修改后的 Excel 文件
        workbook.save(output_excel)


        # 发送进度更新信号，计算进度百分比
        current_step += 1
        progress = int((current_step / total_steps) * 100)
        self.progress_signal.emit(progress)  # 发出进度更新信号
        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"程序结束于: {time.ctime(end_time)}")
        print(f"程序总运行时间: {elapsed_time:.2f}秒")
        # 任务完成后发射 finished_signal 信号
        self.finished_signal.emit()

class DltConverter(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        # 初始化 QSettings
        self.settings = QSettings("ZhaixiyangTools", "DLT2XLSX")
        self.load_settings()

    def initUI(self):
        self.setWindowTitle("DLT文件转换为xlsx-thread版")


        self.setGeometry(500, 200, 500, 600)
        self.label_dlt_file = QLabel("DLT文件夹:", self)
        self.label_dlt_file.move(10, 20)
        self.entry_dlt_file = QLineEdit(self)
        self.entry_dlt_file.setGeometry(105, 20, 300, 25)
        self.button_select_dlt_file = QPushButton("选择文件夹", self)
        self.button_select_dlt_file.setGeometry(412, 20, 80, 25)
        self.button_select_dlt_file.clicked.connect(self.select_dlt_file)
        self.label_example_dlt_file = QLabel("选择DLT文件夹", self)
        self.label_example_dlt_file.move(10, 42)
        self.label_example_dlt_file.setFixedWidth(450)  # 设置标签部件的宽度为 400
        self.label_example_dlt_file.setStyleSheet("color: red; font-family: 'Microsoft YaHei';")

        self.label_dlt_viewer = QLabel("DLT Viewer路径:", self)
        self.label_dlt_viewer.move(10, 75)
        self.entry_dlt_viewer = QLineEdit(self)
        self.entry_dlt_viewer.setGeometry(105, 75, 300, 25)
        self.button_select_dlt_viewer = QPushButton("输入路径", self)
        self.button_select_dlt_viewer.setGeometry(412, 75, 80, 25)
        self.label_example_dlt_viewer = QLabel("示例路径：E:\DltViewer_2.17.0_Stable (也就是dlt_viewer.exe所在的路径)<br>仅需手动输入一次，后续再打开程序默认保留上次的输入", self)
        self.label_example_dlt_viewer.move(10, 100)
        self.label_example_dlt_viewer.setFixedWidth(450)  # 设置标签部件的宽度为 400
        self.label_example_dlt_viewer.setFixedHeight(55)
        self.label_example_dlt_viewer.setStyleSheet("color: red; font-family: 'Microsoft YaHei';")

        self.label_instructions = QLabel("<br>使用说明：<br>1.将要转换的日志文件放入同一个文件夹<br>2.日志文件可以是.gz压缩包，也可以是.dlt，也可以是.log<br>"
                                         "3.转换完成后统统会变为.dlt文件（原始.gz压缩包不会被删除）<br>4.完成后所有日志会被合并进同一个xlsx文件，;-)<br><br>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;如有使用问题请联系<br>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;翟喜洋07900<br>", self)
        self.label_instructions.move(10, 115)
        self.label_instructions.setFixedWidth(450)  # 设置标签部件的宽度为 400
        self.label_instructions.setFixedHeight(200)
        self.label_instructions.setStyleSheet("color: black; font-family: 'Microsoft YaHei';")

        self.button_convert = QPushButton("转♂换", self)
        self.button_convert.setGeometry(200, 300, 100, 30)
        self.button_convert.clicked.connect(self.convert)
        # self.button_convert.setStyleSheet(
        #     """
        #     QPushButton {
        #         background: linear-gradient(135deg, #00FFFF, #0000FF, #8A2BE2, #FF00FF);
        #         border: 2px solid #00FFFF;
        #         border-radius: 15px;
        #         color: black;
        #         font-size: 16px;
        #         font-weight: bold;
        #         text-align: center;
        #         padding: 5px 15px;
        #         text-shadow: 1px 1px 5px rgba(255, 255, 255, 0.6);
        #         box-shadow: 0 0 15px rgba(0, 255, 255, 0.8), 0 0 25px rgba(255, 0, 255, 0.8);
        #         transition: all 0.4s ease-in-out;
        #         outline: none;
        #     }
        #     QPushButton:hover {
        #         background: linear-gradient(135deg, #FF00FF, #FF6347, #FFD700, #00FF7F);
        #         border-color: #FFD700;
        #         transform: scale(1.1);
        #         box-shadow: 0 0 20px rgba(0, 255, 255, 1), 0 0 40px rgba(255, 0, 255, 1);
        #         filter: brightness(1.3);
        #     }
        #     QPushButton:pressed {
        #         background: linear-gradient(135deg, #0000FF, #8A2BE2, #FF1493, #00FFFF);
        #         border-color: #8A2BE2;
        #         transform: scale(1.05);
        #         box-shadow: 0 0 10px rgba(255, 0, 255, 0.9), 0 0 15px rgba(0, 255, 255, 0.9);
        #         filter: brightness(1.1);
        #     }
        #     """
        # )

        self.progress_bar = QProgressBar(self)
        self.progress_bar.setGeometry(11, 340, 480, 20)
        self.progress_bar.setValue(0)
        self.progress_bar.setStyleSheet(
            """
            QProgressBar {
                border: 2px solid #1E90FF;
                border-radius: 10px;
                background-color: #E0E0E0;
                text-align: center;
                color: #1E90FF;
            }
            QProgressBar::chunk {
                background-color: qradialgradient(cx:0.5, cy:0.5, radius:0.5, fx:0.5, fy:0.5, stop:0 rgba(30, 144, 255, 255), stop:1 rgba(30, 144, 255, 0));
                border-radius: 10px;
                border: 2px solid #1E90FF;
            }
            QProgressBar::chunk:disabled {
                background-color: rgba(0, 0, 0, 100);
            }
            """
        )
        self.text_edit = QTextEdit(self)
        self.text_edit.setReadOnly(True)  # 设置为只读，防止用户修改内容
        self.text_edit.setGeometry(11, 375, 480, 200)
        self.text_edit.setStyleSheet("""
                    QTextEdit {
                        background-color: #CCE8CF;  /* 护眼色背景：淡绿色 */
                        border: 2px solid #1e90ff;  /* 蓝色科技范边框 */
                        border-radius: 5px;  /* 圆角边框 */
                        padding: 10px;
                        font-family: '微软雅黑';  /* 设置字体为微软雅黑 */
                        font-size: 14px;  /* 设置字体大小 */
                    }
                """)
        # 将标准输出重定向到 QTextEdit 控件
        sys.stdout = StreamToTextEdit(self.text_edit)

    # def on_stdout_write(self, text):
    #     """这个方法会接收输出的文本，并插入到 QTextEdit 控件中"""
    #     # self.text_edit.append(text)
    #     self.text_edit.insertPlainText(text)
    #     self.text_edit.ensureCursorVisible()  # 确保光标可见，滚动到底部

    def load_settings(self):
        # 从设置中加载已保存的值到文本框
        dlt_viewer_path = self.settings.value("DLTViewerPath", "")
        self.entry_dlt_viewer.setText(dlt_viewer_path)

    def save_settings(self):
        # 将当前文本框中的值保存到设置中
        dlt_viewer_path = self.entry_dlt_viewer.text()
        self.settings.setValue("DLTViewerPath", dlt_viewer_path)

    def closeEvent(self, event):
        self.save_settings()
        event.accept()

    def select_dlt_file(self):
        file_path = QFileDialog.getExistingDirectory(self, "选择DLT文件夹")
        if file_path:
            self.entry_dlt_file.setText(file_path)

    def convert(self):
        folder_path = self.entry_dlt_file.text()
        dlt_viewer_path = self.entry_dlt_viewer.text() + "\dlt_viewer"
        if not os.path.isdir(folder_path):
            QMessageBox.critical(self, "错误", "请选择有效的DLT文件")
            return
        # 禁用按钮，防止多次点击
        self.button_convert.setEnabled(False)
        self.worker_thread = MyWorkerThread(folder_path, dlt_viewer_path)  # 创建后台线程对象
        self.worker_thread.progress_signal.connect(self.update_progress)  # 连接进度更新信号
        self.worker_thread.finished_signal.connect(self.on_conversion_finished)  # 连接任务完成信号
        self.worker_thread.dltviewer_error_signal.connect(lambda: [QMessageBox.warning(self, "小问题", "DltViewer路径出错或为空"), self.button_convert.setEnabled(True)])
        self.worker_thread.start()  # 启动后台线程

    def update_progress(self, progress):
        # 更新进度条的值
        self.progress_bar.setValue(progress)

    def on_conversion_finished(self):
        # 在任务完成后显示提示框
        QMessageBox.information(self, "提示", "转换完成，xlsx文件已保存在dlt文件相同路径的文件夹下。")
        self.button_convert.setEnabled(True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = DltConverter()
    window.show()
    sys.exit(app.exec_())





